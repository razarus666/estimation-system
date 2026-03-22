import os
import uuid
import csv
import json
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime
from functools import wraps
from io import BytesIO

import bcrypt
import pdfplumber
from flask import (
    Flask, render_template, request, redirect, url_for, session, jsonhify, send_file
)
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from models import (
    get_db, init_db, create_admin_user, add_audit_log, add_error_log
)
from matching_engine import run_project_matching, load_master_data

# Initialize Flask app
app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('FLASK_SECRET_KEY', os.getenv('SECRET_KEY', os.urandom(24).hex()))
app.config['UPLOAD_FOLDER'] = os.getenv('UPLOAD_FOLDER', os.path.join(os.path.dirname(__file__), 'uploads'))
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 100MB max file size
ALLOWED_EXTENSIONS = {'pdf', 'xlsx', 'xls', 'csv', 'tsv', 'shd', 'str', 'txt', 'mdb', 'rak'}

# Email configuration
SMTP_SERVER = os.getenv('SMTP_SERVER', 'smtp.gmail.com')
SMTP_PORT = int(os.getenv('SMTP_PORT', '587'))
SMTP_USERNAME = os.getenv('SMTP_USERNAME', '')
SMTP_PASSWORD = os.getenv('SMTP_PASSWORD', '')
SMTP_FROM_EMAIL = os.getenv('SMTP_FROM_EMAIL', os.getenv('SMTP_USERNAME', ''))
SMTP_FROM_NAME = os.getenv('SMTP_FROM_NAME', '電気設備積算システム')
APP_URL = os.getenv('APP_URL', 'https://estimation-system.onrender.com')
ADMIN_EMAIL = os.getenv('ADMIN_EMAIL', '')
ADMIN_CONTACT = os.getenv('ADMIN_CONTACT', '')


# ==================== EMAIL SYSTEM ====================

def _email_header_html():
    """Common email header"""
    return '''<div style="background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); padding: 30px; border-radius: 12px 12px 0 0; text-align: center;">
        <h1 style="color: white; margin: 0; font-size: 22px;">&#9889; 電気設備積算システム</h1>
        <p style="color: rgba(255,255,255,0.8); margin: 5px 0 0; font-size: 13px;">Japanese Electrical Equipment Estimation System</p>
    </div>'''


def _email_footer_html():
    """Common email footer"""
    contact_line = ''
    if ADMIN_CONTACT:
        contact_line = f'<br>※ お問い合わせ: {ADMIN_CONTACT}'
    return f'''<p style="color: #6b7280; font-size: 12px; margin-top: 24px; padding-top: 16px; border-top: 1px solid #e5e7eb;">
        ※ このメールは「電気設備積算システム」から自動送信されています。{contact_line}
        <br>※ 心当たりがない場合は、このメールを無視してください。
    </p>'''


def _email_wrapper(content):
    """Wrap content in email layout"""
    return f'''<div style="font-family: 'Helvetica Neue', Arial, 'Hiragino Sans', 'Hiragino Kaku Gothic ProN', Meiryo, sans-serif; max-width: 600px; margin: 0 auto; padding: 20px; background: #f3f4f6;">
    {_email_header_html()}
    <div style="background: #ffffff; padding: 30px; border: 1px solid #e5e7eb; border-top: none; border-radius: 0 0 12px 12px;">
        {content}
        {_email_footer_html()}
    </div>
</div>'''


def _email_button(url, label, color_from='#10b981', color_to='#059669'):
    """Styled CTA button"""
    return f'''<div style="text-align: center; margin: 28px 0;">
        <a href="{url}" style="background: linear-gradient(135deg, {color_from} 0%, {color_to} 100%); color: white; padding: 14px 32px; border-radius: 8px; text-decoration: none; font-weight: 600; font-size: 15px; display: inline-block;">
            {label}
        </a>
    </div>'''


def _email_info_box(items):
    """Info box with key-value pairs"""
    rows = ''
    for key, val in items:
        rows += f'''<tr>
            <td style="padding: 8px 12px; color: #6b7280; font-size: 13px; white-space: nowrap; vertical-align: top;">{key}</td>
            <td style="padding: 8px 12px; color: #1f2937; font-size: 13px; font-weight: 500;">{val}</td>
        </tr>'''
    return f'''<div style="background: #f9fafb; border: 1px solid #e5e7eb; border-radius: 8px; padding: 4px; margin: 20px 0;">
        <table style="width: 100%; border-collapse: collapse;">{rows}</table>
    </div>'''


def _strip_html(html_body):
    """Simple HTML to plain text for multipart email"""
    import re
    text = re.sub(r'<br\s*/?>', '\n', html_body)
    text = re.sub(r'<[^>]+>', '', text)
    text = re.sub(r'\n\s*\n', '\n\n', text)
    return text.strip()


def log_email(to_email, to_name, subject, email_type, status, error_message=None, triggered_by=None, related_user_id=None):
    """Log email send attempt to database"""
    try:
        db = get_db()
        db.execute(
            '''INSERT INTO email_log (to_email, to_name, subject, email_type, status, error_message, triggered_by, related_user_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (to_email, to_name, subject, email_type, status, error_message, triggered_by, related_user_id)
        )
        db.commit()
    except Exception as e:
        app.logger.error(f'メール送信ログ記録エラー: {str(e)}')


def send_notification_email(to_email, to_name, subject, html_body, email_type='general', triggered_by=None, related_user_id=None):
    """Send email notification using SMTP with logging and duplicate prevention"""
    if not SMTP_USERNAME or not SMTP_PASSWORD:
        app.logger.warning(f'SMTP未設定のためメール送信スキップ: {to_email} [{email_type}]')
        log_email(to_email, to_name, subject, email_type, 'skipped', 'SMTP未設定', triggered_by, related_user_id)
        return False

    try:
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = f'{SMTP_FROM_NAME} <{SMTP_FROM_EMAIL}>'
        msg['To'] = f'{to_name} <{to_email}>'
        msg['X-Mailer'] = 'EstimationSystem/1.3'

        # Attach plain text version first, then HTML
        plain_text = _strip_html(html_body)
        msg.attach(MIMEText(plain_text, 'plain', 'utf-8'))
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))

        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT, timeout=15) as server:
            
            server.login(SMTP_USERNAME, SMTP_PASSWORD)
            server.send_message(msg)

        app.logger.info(f'メール送信成功: {to_email} [{email_type}]')
        log_email(to_email, to_name, subject, email_type, 'sent', None, triggered_by, related_user_id)
        return True

    except Exception as e:
        error_msg = str(e)
        app.logger.error(f'メール送信エラー: {to_email} [{email_type}] - {error_msg}')
        log_email(to_email, to_name, subject, email_type, 'failed', error_msg, triggered_by, related_user_id)
        return False


# --- Email Type 1: Registration → Admin Notification ---
def send_registration_admin_notify(user_name, user_email, registered_at, triggered_by=None, related_user_id=None):
    """Notify admin that a new user has registered and is awaiting approval"""
    if not ADMIN_EMAIL:
        app.logger.warning('ADMIN_EMAIL未設定のため管理者通知スキップ')
        return False

    subject = '【電気設備積算システム】新規ユーザー登録通知（承認待ち）'
    content = f'''
        <h2 style="color: #1a1f36; margin-top: 0; font-size: 18px;">新規ユーザー登録通知</h2>
        <p style="color: #4b5563; line-height: 1.6;">
            新しいユーザーが登録しました。<br>承認待ちのため、管理画面から承認または却下をお願いいたします。
        </p>
        {_email_info_box([
            ('登録者氏名', user_name),
            ('メールアドレス', user_email),
            ('登録日時', registered_at),
            ('ステータス', '<span style="color: #d97706; font-weight: 700;">&#9203; 承認待ち</span>'),
        ])}
        {_email_button(APP_URL + '/admin/users', '管理画面で承認する', '#6366f1', '#4f46e5')}
        <p style="color: #6b7280; font-size: 13px;">承認すると、登録者にログイン可能の通知メールが自動送信されます。</p>
    '''
    html_body = _email_wrapper(content)
    return send_notification_email(ADMIN_EMAIL, '管理者', subject, html_body, 'registration_admin_notify', triggered_by, related_user_id)


# --- Email Type 2: Registration → User Confirmation ---
def send_registration_user_confirm(user_email, user_name, triggered_by=None, related_user_id=None):
    """Send registration confirmation to the user"""
    subject = '【電気設備積算システム】ユーザー登録を受け付けました'
    content = f'''
        <h2 style="color: #1a1f36; margin-top: 0; font-size: 18px;">ユーザー登録を受け付けました</h2>
        <p style="color: #4b5563; line-height: 1.6;">
            {user_name} 様<br><br>
            電気設備積算システムへの登録ありがとうございます。<br>
            ご登録を受け付けいたしました。
        </p>
        <div style="background: #fffbeb; border: 1px solid #fcd34d; border-radius: 8px; padding: 16px; margin: 20px 0;">
            <p style="color: #92400e; margin: 0; font-size: 14px;">
                <strong>&#9888;&#65039; 現在、管理者による承認待ちです</strong><br>
                承認が完了次第、ログイン可能になります。<br>
                承認完了時にメールでお知らせいたしますので、今しばらくお待ちください。
            </p>
        </div>
        {_email_info_box([
            ('登録メールアドレス', user_email),
            ('登録者氏名', user_name),
            ('ログインURL', f'<a href="{APP_URL}/login" style="color: #6366f1;">{APP_URL}/login</a>'),
        ])}
        <p style="color: #6b7280; font-size: 13px;">
            ※ 承認には通常1営業日程度いただいております。<br>
            ※ ご不明な点がございましたら管理者にお問い合わせください。
        </p>
    '''
    html_body = _email_wrapper(content)
    return send_notification_email(user_email, user_name, subject, html_body, 'registration_user_confirm', triggered_by, related_user_id)


# --- Email Type 3: Approval → User Notification ---
def send_approval_email(user_email, user_name, triggered_by=None, related_user_id=None):
    """Send approval notification to the user"""
    subject = '【電気設備積算システム】アカウントが承認されました'
    content = f'''
        <h2 style="color: #1a1f36; margin-top: 0; font-size: 18px;">
            <span style="color: #10b981;">&#10004;</span> アカウントが承認されました
        </h2>
        <p style="color: #4b5563; line-height: 1.6;">
            {user_name} 様<br><br>
            管理者によりアカウントが承認されました。<br>
            以下のボタンからログインして、電気設備積算システムをご利用いただけます。
        </p>
        {_email_button(APP_URL + '/login', 'ログインする')}
        <div style="background: #f0fdf4; border: 1px solid #86efac; border-radius: 8px; padding: 16px; margin: 20px 0;">
            <p style="color: #166534; margin: 0; font-size: 14px;">
                <strong>&#128274; 初回ログイン時のご案内</strong><br>
                ・ 登録時に設定したメールアドレスとパスワードでログインしてください<br>
                ・ セキュリティのため、初回ログイン後にパスワードの変更を推奨します<br>
                ・ ログイン後、ダッシュボードからプロジェクトの作成・見積作成が可能です
            </p>
        </div>
        {_email_info_box([
            ('ログインURL', f'<a href="{APP_URL}/login" style="color: #6366f1;">{APP_URL}/login</a>'),
            ('ログインID', user_email),
        ])}
    '''
    html_body = _email_wrapper(content)
    return send_notification_email(user_email, user_name, subject, html_body, 'approval_notify', triggered_by, related_user_id)


# --- Email Type 4: Rejection → User Notification ---
def send_rejection_email(user_email, user_name, reason='', triggered_by=None, related_user_id=None):
    """Send rejection notification to the user"""
    subject = '【電気設備積算システム】ユーザー登録について'
    reason_html = ''
    if reason:
        reason_html = f'''
        <div style="background: #fef2f2; border: 1px solid #fca5a5; border-radius: 8px; padding: 16px; margin: 20px 0;">
            <p style="color: #991b1b; margin: 0; font-size: 14px;">
                <strong>管理者からのメッセージ:</strong><br>{reason}
            </p>
        </div>'''

    content = f'''
        <h2 style="color: #1a1f36; margin-top: 0; font-size: 18px;">ユーザー登録について</h2>
        <p style="color: #4b5563; line-height: 1.6;">
            {user_name} 様<br><br>
            この度は電気設備積算システムへの登録申請をいただき、誠にありがとうございました。<br>
            大変恐れ入りますが、審査の結果、今回のご登録を承認することができませんでした。
        </p>
        {reason_html}
        <p style="color: #4b5563; line-height: 1.6;">
            ご不明な点やご質問がございましたら、管理者までお気軽にお問い合わせください。<br>
            今後ともよろしくお願いいたします。
        </p>
    '''
    html_body = _email_wrapper(content)
    return send_notification_email(user_email, user_name, subject, html_body, 'rejection_notify', triggered_by, related_user_id)


# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# Create upload folder
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Initialize database (startup handles seed DB copy on Render)
from startup import ensure_db
ensure_db()


class User(UserMixin):
    """User class for Flask-Login"""
    def __init__(self, user_id, email, full_name, role, active):
        self.id = user_id
        self.email = email
        self.full_name = full_name
        self.role = role
        self._active = bool(active)

    @property
    def is_active(self):
        return self._active

    def is_admin(self):
        return self.role == 'admin'

    def is_approved(self):
        return self.role != 'pending'


@login_manager.user_loader
def load_user(user_id):
    """Load user from database"""
    db = get_db()
    cursor = db.cursor()
    cursor.execute(
        'SELECT id, email, full_name, role, is_active FROM users WHERE id = ?',
        (user_id,)
    )
    user_data = cursor.fetchone()
    if user_data:
        return User(user_data[0], user_data[1], user_data[2], user_data[3], user_data[4])
    return None


def admin_required(f):
    """Decorator for admin-only routes"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user or not current_user.is_authenticated:
            return redirect(url_for('login'))
        if not current_user.is_admin():
            add_audit_log(
                current_user.id,
                'UNAUTHORIZED_ACCESS',
                'route',
                request.path,
                'WARNING',
                f'Unauthorized admin access attempt: {request.path}',
                request.remote_addr
            )
            return render_template('error.html', error_code=403, error_message='管理者権限が必要です'), 403
        return f(*args, **kwargs)
    return decorated_function


def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_user_ip():
    """Get client IP address"""
    return request.headers.get('X-Forwarded-For', request.remote_addr).split(',')[0].strip()


def extract_pdf_text(file_path):
    """Extract text from PDF using pdfplumber"""
    try:
        text_content = []
        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                text_content.append({
                    'page': page_num,
                    'text': text
                })
        return text_content
    except Exception as e:
        raise Exception(f'PDF解析エラー: {str(e)}')


def _normalize_header(text):
    """Normalize header text by removing full-width/half-width spaces for comparison"""
    import re
    if not text:
        return ''
    # Remove all whitespace (full-width space \u3000, half-width space, tabs, etc.)
    return re.sub(r'[\s\u3000]+', '', str(text).strip()).lower()


def parse_material_list_excel(file_path):
    """Parse material list from Excel file"""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)

        # Header keywords to search for (normalized, no spaces)
        header_keywords = ['名称', 'name', 'material', '品名', '材料名']

        # Try all sheets (prioritize sheets with names like '内訳', 'material', 'data')
        priority_sheet_names = ['内訳', '明細', 'material', 'data', '材料']
        sheets_to_try = []

        # Add priority sheets first
        for sn in priority_sheet_names:
            for ws_name in wb.sheetnames:
                if sn in ws_name.lower() or sn in _normalize_header(ws_name):
                    if ws_name not in sheets_to_try:
                        sheets_to_try.append(ws_name)

        # Then add remaining sheets
        for ws_name in wb.sheetnames:
            if ws_name not in sheets_to_try:
                sheets_to_try.append(ws_name)

        header_row = None
        ws = None

        for sheet_name in sheets_to_try:
            ws_candidate = wb[sheet_name]
            for row_idx, row in enumerate(ws_candidate.iter_rows(values_only=True), 1):
                if row_idx > 50:  # Don't search beyond row 50
                    break
                row_normalized = [_normalize_header(v) for v in row]
                if any(kw in cell for cell in row_normalized for kw in header_keywords):
                    header_row = row_idx
                    ws = ws_candidate
                    break
            if header_row:
                break

        if not header_row or not ws:
            raise Exception('ヘッダー行が見つかりません（全シート検索済み）')

        # Get headers (normalize for matching but keep original for display)
        headers = []
        for cell in ws[header_row]:
            headers.append(str(cell.value).strip() if cell.value else '')

        # Expected columns (Japanese) — keys are normalized (no spaces)
        column_mapping = {
            '行番号': 'row_no',
            '名称': 'material_name',
            '品名': 'material_name',
            '規格': 'spec',
            'サイズ': 'size',
            '数量': 'quantity',
            '単位': 'unit',
            '施工条件': 'construction_method',
            '分野': 'field_category',
            '図面参照': 'drawing_ref',
            '備考': 'remarks',
            '単価': 'unit_price',
            '金額': 'amount',
        }

        # Map columns — normalize header text before matching
        column_indices = {}
        for col_idx, header in enumerate(headers):
            header_norm = _normalize_header(header)
            for jp_name, py_name in column_mapping.items():
                jp_norm = _normalize_header(jp_name)
                if jp_norm in header_norm or py_name in header_norm:
                    if py_name not in column_indices:  # Don't overwrite first match
                        column_indices[py_name] = col_idx
                    break

        # Parse rows
        materials = []
        row_counter = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_idx <= header_row:
                continue
            if all(cell is None for cell in row):
                continue

            material = {}
            for py_name, col_idx in column_indices.items():
                if col_idx < len(row):
                    value = row[col_idx]
                    if py_name in ('quantity', 'unit_price', 'amount'):
                        try:
                            material[py_name] = float(value) if value else 0
                        except:
                            material[py_name] = 0
                    else:
                        material[py_name] = str(value).strip() if value else ''
                else:
                    material[py_name] = '' if py_name not in ('quantity', 'unit_price', 'amount') else 0

            # Skip section headers (rows like "小　　　計", "合　　計") and empty names
            name = material.get('material_name', '')
            if name and '計' not in _normalize_header(name):
                row_counter += 1
                if 'row_no' not in material or not material['row_no']:
                    material['row_no'] = str(row_counter)
                materials.append(material)

        return materials
    except Exception as e:
        raise Exception(f'Excel解析エラー: {str(e)}')


def parse_material_list_csv(file_path):
    """Parse material list from CSV file"""
    try:
        materials = []
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames:
                raise Exception('CSVが空です')

            column_mapping = {
                '行番号': 'row_no',
                '名称': 'material_name',
                '規格': 'spec',
                'サイズ': 'size',
                '数量': 'quantity',
                '単位': 'unit',
                '施工条件': 'construction_method',
                '分野': 'field_category',
                '図面参照': 'drawing_ref',
                '備考': 'remarks'
            }

            for row in reader:
                material = {}
                for py_name in column_mapping.values():
                    material[py_name] = ''

                for jp_name, py_name in column_mapping.items():
                    for csv_col in reader.fieldnames:
                        if jp_name.lower() in csv_col.lower() or py_name in csv_col.lower():
                            value = row.get(csv_col, '')
                            if py_name == 'quantity':
                                try:
                                    material[py_name] = float(value) if value else 0
                                except:
                                    material[py_name] = 0
                            else:
                                material[py_name] = str(value).strip() if value else ''
                            break

                if material.get('material_name'):
                    materials.append(material)

        return materials
    except Exception as e:
        raise Exception(f'CSV解析エラー: {str(e)}')


def parse_material_list_shd(file_path):
    """Parse material list from Adonis IXAS .shd file (Shift-JIS encoded)"""
    try:
        materials = []
        with open(file_path, 'rb') as f:
            # Read as Shift-JIS with error handling
            content = f.read().decode('shift_jis', errors='replace')

        # Split by carriage return and then by tab
        lines = content.split('\r')

        for row_idx, line in enumerate(lines):
            fields = line.split('\t')

            # Skip header row (row 0)
            if row_idx == 0:
                continue

            # Skip if no fields or code (col 0/1) is empty
            if not fields or (len(fields) > 1 and not fields[1].strip()):
                continue

            # Skip category headers (name starts with ◆ or 【)
            if len(fields) > 2 and fields[2]:
                name = fields[2].strip()
                if name.startswith('◆') or name.startswith('【'):
                    continue

            # Extract fields (adjust indices based on .shd format)
            # Typical format: col0=binary, col1=code, col2=name, col3=spec, col4=construction_method, col5=quantity, col6=unit, col7=unit_price
            if len(fields) >= 8:
                material = {
                    'row_no': row_idx,
                    'material_name': fields[2].strip() if len(fields) > 2 else '',
                    'spec': fields[3].strip() if len(fields) > 3 else '',
                    'construction_method': fields[4].strip() if len(fields) > 4 else '',
                    'quantity': 0,
                    'unit': fields[6].strip() if len(fields) > 6 else '',
                    'field_category': '',
                    'size': '',
                    'drawing_ref': '',
                    'remarks': ''
                }

                # Try to parse quantity and unit_price
                if len(fields) > 5:
                    try:
                        material['quantity'] = float(fields[5].strip()) if fields[5].strip() else 0
                    except:
                        material['quantity'] = 0

                if material.get('material_name'):
                    materials.append(material)

        return materials
    except Exception as e:
        raise Exception(f'SHDファイル解析エラー: {str(e)}')


# ==================== AUTH ROUTES ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login page"""
    if request.method == 'POST':
        try:
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '').strip()

            if not email or not password:
                return render_template('login.html', error='メールアドレスとパスワードを入力してください'), 400

            db = get_db()
            cursor = db.cursor()
            cursor.execute(
                'SELECT id, email, password_hash, full_name, role, is_active FROM users WHERE email = ?',
                (email,)
            )
            user_data = cursor.fetchone()

            if not user_data:
                add_audit_log(
                    None,
                    'LOGIN_FAILED',
                    'user',
                    email,
                    'WARNING',
                    'ユーザーが見つかりません',
                    get_user_ip()
                )
                return render_template('login.html', error='メールアドレスまたはパスワードが正しくありません'), 401

            user_id, db_email, password_hash, full_name, role, is_active = user_data

            if not is_active:
                add_audit_log(
                    user_id,
                    'LOGIN_FAILED',
                    'user',
                    email,
                    'WARNING',
                    'アカウントが無効化されています',
                    get_user_ip()
                )
                return render_template('login.html', error='アカウントが無効化されています'), 403

            pw_hash_bytes = password_hash.encode('utf-8') if isinstance(password_hash, str) else password_hash
            if not bcrypt.checkpw(password.encode('utf-8'), pw_hash_bytes):
                add_audit_log(
                    user_id,
                    'LOGIN_FAILED',
                    'user',
                    email,
                    'WARNING',
                    'パスワード不正',
                    get_user_ip()
                )
                return render_template('login.html', error='メールアドレスまたはパスワードが正しくありません'), 401

            user = User(user_id, db_email, full_name, role, is_active)
            login_user(user)

            # Record last login time
            cursor.execute(
                'UPDATE users SET last_login_at = ? WHERE id = ?',
                (datetime.utcnow(), user_id)
            )
            db.commit()

            add_audit_log(
                user_id,
                'LOGIN',
                'user',
                email,
                'INFO',
                'ログイン成功',
                get_user_ip()
            )

            # Redirect to pending page if user is pending approval
            if role == 'pending':
                return redirect(url_for('pending_page'))

            return redirect(url_for('dashboard'))

        except Exception as e:
            add_error_log(
                current_user.id if current_user.is_authenticated else None,
                'LOGIN_ERROR',
                str(e),
                str(e),
                request.url
            )
            return render_template('login.html', error='ログイン処理中にエラーが発生しました'), 500

    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """Register new user"""
    if request.method == 'POST':
        try:
            email = request.form.get('email', '').strip()
            password = request.form.get('password', '').strip()
            full_name = request.form.get('full_name', '').strip()

            if not all([email, password, full_name]):
                return render_template('register.html', error='すべてのフィールドを入力してください'), 400

            if len(password) < 8:
                return render_template('register.html', error='パスワードは8文字以上である必要があります'), 400

            db = get_db()
            cursor = db.cursor()

            cursor.execute('SELECT id FROM users WHERE email = ?', (email,))
            if cursor.fetchone():
                return render_template('register.html', error='このメールアドレスは既に登録されています'), 400

            password_hash = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())

            cursor.execute(
                '''INSERT INTO users (email, password_hash, full_name, role, is_active, created_at)
                   VALUES (?, ?, ?, ?, ?, ?)''',
                (email, password_hash, full_name, 'pending', True, datetime.utcnow())
            )
            db.commit()

            new_user_id = cursor.lastrowid
            registered_at = datetime.utcnow().strftime('%Y-%m-%d %H:%M')

            add_audit_log(
                new_user_id,
                'REGISTER',
                'user',
                email,
                'INFO',
                f'ユーザー登録完了（承認待ち）',
                get_user_ip()
            )

            # Send registration confirmation email to user
            email_user_ok = send_registration_user_confirm(email, full_name, triggered_by=None, related_user_id=new_user_id)

            # Send notification email to admin
            email_admin_ok = send_registration_admin_notify(full_name, email, registered_at, triggered_by=None, related_user_id=new_user_id)

            add_audit_log(
                new_user_id,
                'REGISTRATION_EMAIL',
                'email',
                email,
                'INFO',
                f'登録メール送信 (ユーザー: {"成功" if email_user_ok else "未送信"}, 管理者: {"成功" if email_admin_ok else "未送信"})',
                get_user_ip()
            )

            return render_template(
                'register.html',
                success='登録完了しました。管理者の承認をお待ちください。確認メールをお送りしましたのでご確認ください。'
            )

        except Exception as e:
            add_error_log(
                current_user.id if current_user.is_authenticated else None,
                'REGISTER_ERROR',
                str(e),
                str(e),
                request.url
            )
            return render_template('register.html', error='登録処理中にエラーが発生しました'), 500

    return render_template('register.html')


@app.route('/logout')
@login_required
def logout():
    """Logout"""
    add_audit_log(
        current_user.id,
        'LOGOUT',
        'user',
        current_user.email,
        'INFO',
        'ログアウト',
        get_user_ip()
    )
    logout_user()
    return redirect(url_for('login'))


# ==================== USER PROFILE ROUTES ====================

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    """User profile page"""
    try:
        db = get_db()
        cursor = db.cursor()

        if request.method == 'POST':
            # Update profile
            full_name = request.form.get('full_name', '').strip()
            phone = request.form.get('phone', '').strip()
            department = request.form.get('department', '').strip()

            cursor.execute(
                '''UPDATE users
                   SET full_name = ?, phone = ?, department = ?
                   WHERE id = ?''',
                (full_name, phone, department, current_user.id)
            )
            db.commit()

            add_audit_log(
                current_user.id,
                'UPDATE_PROFILE',
                'user',
                current_user.id,
                'INFO',
                'プロフィール更新',
                get_user_ip()
            )

            # Reload user object
            user = load_user(current_user.id)
            login_user(user)

            return render_template('profile.html', user=user, message='プロフィールを更新しました')

        # GET: Show profile page
        cursor.execute(
            'SELECT id, email, full_name, phone, department, avatar_path FROM users WHERE id = ?',
            (current_user.id,)
        )
        user_data = cursor.fetchone()

        return render_template('profile.html', user=user_data)

    except Exception as e:
        add_error_log(
            current_user.id,
            'PROFILE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='プロフィールの読み込みに失敗しました'), 500


@app.route('/profile/avatar', methods=['POST'])
@login_required
def upload_avatar():
    """Upload user avatar"""
    try:
        if 'avatar' not in request.files:
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        file = request.files['avatar']
        if file.filename == '':
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        # Check file type
        if not file.filename.lower().endswith(('.png', '.jpg', '.jpeg', '.gif')):
            return jsonify({'error': '画像ファイル（PNG, JPG, GIF）のみサポートしています'}), 400

        # Create user avatar directory
        avatar_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'avatars')
        os.makedirs(avatar_dir, exist_ok=True)

        # Save avatar
        file_ext = file.filename.rsplit('.', 1)[1].lower()
        avatar_filename = f"avatar_{current_user.id}.{file_ext}"
        avatar_path = os.path.join(avatar_dir, avatar_filename)
        file.save(avatar_path)

        # Update database
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            'UPDATE users SET avatar_path = ? WHERE id = ?',
            (f'avatars/{avatar_filename}', current_user.id)
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'UPLOAD_AVATAR',
            'user',
            current_user.id,
            'INFO',
            'アバター画像アップロード',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'アバター画像をアップロードしました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'UPLOAD_AVATAR_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'アバターアップロードエラー: {str(e)}'}), 500


# ==================== PENDING PAGE ROUTE ====================

@app.route('/pending')
@login_required
def pending_page():
    """Pending approval page"""
    try:
        add_audit_log(
            current_user.id,
            'VIEW_PENDING_PAGE',
            'user',
            current_user.id,
            'INFO',
            '承認待ちページ表示',
            get_user_ip()
        )
        return render_template('pending.html')
    except Exception as e:
        add_error_log(
            current_user.id,
            'PENDING_PAGE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='承認待ちページの読み込みに失敗しました'), 500


# ==================== DASHBOARD ROUTES ====================

@app.route('/')
def index():
    """Redirect to dashboard"""
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard - show projects list and stats"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Get projects for this user
        cursor.execute(
            '''SELECT id, name, description, client_name, status, created_at, updated_at
               FROM projects
               WHERE created_by = ?
               ORDER BY updated_at DESC''',
            (current_user.id,)
        )
        projects = cursor.fetchall()

        project_list = []
        for project in projects:
            project_list.append({
                'id': project[0],
                'name': project[1],
                'description': project[2],
                'client_name': project[3],
                'status': project[4],
                'created_at': project[5],
                'updated_at': project[6]
            })

        # Compute dashboard statistics
        # Total projects
        total_projects = len(project_list)

        # Total estimate details
        cursor.execute(
            'SELECT COUNT(*) FROM estimate_details WHERE project_id IN (SELECT id FROM projects WHERE created_by = ?)',
            (current_user.id,)
        )
        total_estimates = cursor.fetchone()[0] or 0

        # Estimates needing review (confidence < 0.75)
        cursor.execute(
            '''SELECT COUNT(*) FROM estimate_details ed
               INNER JOIN projects p ON ed.project_id = p.id
               WHERE p.created_by = ? AND ed.confidence < 0.75''',
            (current_user.id,)
        )
        needs_review = cursor.fetchone()[0] or 0

        # Total amount (sum of all estimate amounts)
        cursor.execute(
            '''SELECT COALESCE(SUM(amount), 0) FROM estimate_details ed
               INNER JOIN projects p ON ed.project_id = p.id
               WHERE p.created_by = ?''',
            (current_user.id,)
        )
        total_amount = cursor.fetchone()[0] or 0

        # Total productivity
        cursor.execute(
            '''SELECT COALESCE(SUM(productivity_total), 0) FROM estimate_details ed
               INNER JOIN projects p ON ed.project_id = p.id
               WHERE p.created_by = ?''',
            (current_user.id,)
        )
        total_productivity = cursor.fetchone()[0] or 0

        # Error count (last 7 days)
        cursor.execute(
            '''SELECT COUNT(*) FROM error_log
               WHERE user_id = ? AND datetime(created_at) > datetime('now', '-7 days')''',
            (current_user.id,)
        )
        error_count = cursor.fetchone()[0] or 0

        stats = {
            'total_projects': total_projects,
            'total_estimates': total_estimates,
            'needs_review': needs_review,
            'total_amount': total_amount,
            'total_productivity': total_productivity,
            'error_count': error_count
        }

        add_audit_log(
            current_user.id,
            'VIEW_DASHBOARD',
            'dashboard',
            None,
            'INFO',
            'ダッシュボード表示',
            get_user_ip()
        )

        return render_template('dashboard.html', projects=project_list, stats=stats)

    except Exception as e:
        add_error_log(
            current_user.id,
            'DASHBOARD_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='ダッシュボード読み込みエラーが発生しました'), 500


# ==================== PROJECT ROUTES ====================

@app.route('/projects/new', methods=['GET', 'POST'])
@login_required
def create_project():
    """Create new project"""
    if request.method == 'POST':
        try:
            name = request.form.get('project_name', '').strip()
            description = request.form.get('description', '').strip()
            client_name = request.form.get('client_name', '').strip()
            location = request.form.get('location', '').strip()

            if not name:
                return render_template('project_new.html', error='プロジェクト名は必須です'), 400

            db = get_db()
            cursor = db.cursor()

            cursor.execute(
                '''INSERT INTO projects (name, description, client_name, location, created_by, created_at, updated_at, status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
                (name, description, client_name, location, current_user.id, datetime.utcnow(), datetime.utcnow(), 'draft')
            )
            db.commit()

            project_id = cursor.lastrowid

            add_audit_log(
                current_user.id,
                'CREATE_PROJECT',
                'project',
                project_id,
                'INFO',
                f'プロジェクト作成: {name}',
                get_user_ip()
            )

            return redirect(url_for('project_detail', project_id=project_id))

        except Exception as e:
            add_error_log(
                current_user.id,
                'CREATE_PROJECT_ERROR',
                str(e),
                str(e),
                request.url
            )
            return render_template('project_new.html', error='プロジェクト作成エラーが発生しました'), 500

    return render_template('project_new.html')


@app.route('/projects/<int:project_id>/edit', methods=['POST'])
@login_required
def edit_project(project_id):
    """Edit project details"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        # Get request data
        data = request.get_json() or request.form

        name = data.get('name', '').strip()
        description = data.get('description', '').strip()
        client_name = data.get('client_name', '').strip()
        location = data.get('location', '').strip()

        if not name:
            return jsonify({'error': 'プロジェクト名は必須です'}), 400

        # Update project
        cursor.execute(
            '''UPDATE projects
               SET name = ?, description = ?, client_name = ?, location = ?, updated_at = ?
               WHERE id = ?''',
            (name, description, client_name, location, datetime.utcnow(), project_id)
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'EDIT_PROJECT',
            'project',
            project_id,
            'INFO',
            f'プロジェクト編集: {name}',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'プロジェクトを更新しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'EDIT_PROJECT_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'プロジェクト編集エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/delete', methods=['POST'])
@login_required
def delete_project(project_id):
    """Delete project and all related data"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        # Delete related data
        cursor.execute('DELETE FROM edit_history WHERE project_id = ?', (project_id,))
        cursor.execute('DELETE FROM estimate_details WHERE project_id = ?', (project_id,))
        cursor.execute('DELETE FROM match_results WHERE project_id = ?', (project_id,))
        cursor.execute('DELETE FROM material_list WHERE project_id = ?', (project_id,))
        cursor.execute('DELETE FROM project_files WHERE project_id = ?', (project_id,))
        cursor.execute('DELETE FROM projects WHERE id = ?', (project_id,))

        db.commit()

        # Delete uploaded files from disk
        project_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(project_id))
        if os.path.exists(project_upload_dir):
            import shutil
            shutil.rmtree(project_upload_dir)

        add_audit_log(
            current_user.id,
            'DELETE_PROJECT',
            'project',
            project_id,
            'INFO',
            f'プロジェクト削除: ID {project_id}',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'プロジェクトを削除しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'DELETE_PROJECT_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'プロジェクト削除エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>')
@login_required
def project_detail(project_id):
    """Project detail page"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute(
            'SELECT id, name, description, client_name, status, created_at, created_by FROM projects WHERE id = ?',
            (project_id,)
        )
        project = cursor.fetchone()

        if not project:
            return render_template('error.html', error_code=404, error_message='プロジェクトが見つかりません'), 404

        if project[6] != current_user.id and not current_user.is_admin():
            return render_template('error.html', error_code=403, error_message='アクセス権限がありません'), 403

        # Get project files
        cursor.execute(
            '''SELECT id, file_type, original_name, file_size, uploaded_at
               FROM project_files
               WHERE project_id = ?
               ORDER BY uploaded_at DESC''',
            (project_id,)
        )
        files = cursor.fetchall()

        # Get material list
        cursor.execute(
            '''SELECT id, row_no, material_name, spec, size, quantity, unit, construction_method, field_category
               FROM material_list
               WHERE project_id = ?
               ORDER BY row_no''',
            (project_id,)
        )
        materials = cursor.fetchall()

        # Get estimate details count and totals
        cursor.execute(
            'SELECT COUNT(*), COALESCE(SUM(amount),0), COALESCE(SUM(productivity_total),0) FROM estimate_details WHERE project_id = ?',
            (project_id,)
        )
        est_row = cursor.fetchone()
        estimate_count = est_row[0]
        total_amount = est_row[1]
        total_productivity = est_row[2]

        # Get labor unit price setting
        cursor.execute("SELECT setting_value FROM estimate_settings WHERE setting_key='labor_unit_price'")
        lup_row = cursor.fetchone()
        labor_unit_price = float(lup_row[0]) if lup_row else 25000

        # Get match results for display (join with material_list for material_name)
        cursor.execute(
            '''SELECT mr.id, mr.material_id, mr.candidate_rank, mr.master_id, mr.match_type,
               mr.confidence, mr.reason, mr.is_adopted, mr.master_name, mr.master_spec,
               mr.master_method, mr.composite_unit_price, mr.removal_productivity, mr.source_page,
               COALESCE(ml.material_name, '') as material_name
            FROM match_results mr
            LEFT JOIN material_list ml ON mr.material_id = ml.id
            WHERE mr.project_id = ? ORDER BY mr.material_id, mr.candidate_rank''',
            (project_id,)
        )
        match_results = cursor.fetchall()

        add_audit_log(
            current_user.id,
            'VIEW_PROJECT',
            'project',
            str(project_id),
            'INFO',
            'プロジェクト詳細表示',
            get_user_ip()
        )

        return render_template(
            'project_detail.html',
            project={
                'id': project[0],
                'name': project[1],
                'description': project[2],
                'client_name': project[3],
                'status': project[4],
                'created_at': project[5]
            },
            files=files,
            materials=materials,
            estimate_count=estimate_count,
            total_amount=total_amount,
            total_productivity=total_productivity,
            labor_unit_price=labor_unit_price,
            match_results=match_results
        )

    except Exception as e:
        add_error_log(
            current_user.id,
            'PROJECT_DETAIL_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='プロジェクト詳細読み込みエラーが発生しました'), 500


@app.route('/projects/<int:project_id>/upload', methods=['POST'])
@login_required
def upload_file(project_id):
    """Upload project file"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        if 'file' not in request.files:
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        if not allowed_file(file.filename):
            return jsonify({'error': '許可されないファイル形式です（PDF, Excel, CSV, TSV, SHD, STR, TXT, MDB, RAK）'}), 400

        # Create project upload directory
        project_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], str(project_id))
        os.makedirs(project_upload_dir, exist_ok=True)

        # Save file — secure_filename strips non-ASCII chars (e.g. Japanese),
        # so extract the extension from the original filename first
        original_name = file.filename
        file_ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''
        if not file_ext:
            return jsonify({'error': 'ファイル拡張子を判別できません'}), 400
        filename = secure_filename(file.filename) or f"upload_{uuid.uuid4().hex[:8]}.{file_ext}"
        if '.' not in filename:
            filename = f"{filename}.{file_ext}"
        unique_filename = f"{uuid.uuid4().hex}.{file_ext}"
        file_path = os.path.join(project_upload_dir, unique_filename)
        file.save(file_path)

        file_size = os.path.getsize(file_path)

        # Determine file type and process
        if file_ext == 'pdf':
            file_type = 'estimate_pdf'
            # Extract PDF text
            text_content = extract_pdf_text(file_path)
            stored_path = file_path
        elif file_ext == 'xlsx':
            file_type = 'material_list'
            # Parse Excel and import materials
            materials = parse_material_list_excel(file_path)
            for material in materials:
                cursor.execute(
                    '''INSERT INTO material_list
                       (project_id, row_no, material_name, spec, size, quantity, unit, construction_method, field_category, drawing_ref, remarks)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        project_id,
                        material.get('row_no', ''),
                        material.get('material_name', ''),
                        material.get('spec', ''),
                        material.get('size', ''),
                        material.get('quantity', 0),
                        material.get('unit', ''),
                        material.get('construction_method', ''),
                        material.get('field_category', ''),
                        material.get('drawing_ref', ''),
                        material.get('remarks', '')
                    )
                )
            db.commit()
            stored_path = file_path
        elif file_ext in ('csv', 'tsv', 'txt'):
            file_type = 'material_list'
            # Parse CSV/TSV and import materials
            materials = parse_material_list_csv(file_path)
            for material in materials:
                cursor.execute(
                    '''INSERT INTO material_list
                       (project_id, row_no, material_name, spec, size, quantity, unit, construction_method, field_category, drawing_ref, remarks)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        project_id,
                        material.get('row_no', ''),
                        material.get('material_name', ''),
                        material.get('spec', ''),
                        material.get('size', ''),
                        material.get('quantity', 0),
                        material.get('unit', ''),
                        material.get('construction_method', ''),
                        material.get('field_category', ''),
                        material.get('drawing_ref', ''),
                        material.get('remarks', '')
                    )
                )
            db.commit()
            stored_path = file_path
        elif file_ext == 'shd':
            file_type = 'material_list'
            # Parse SHD (Adonis IXAS) and import materials
            materials = parse_material_list_shd(file_path)
            for material in materials:
                cursor.execute(
                    '''INSERT INTO material_list
                       (project_id, row_no, material_name, spec, size, quantity, unit, construction_method, field_category, drawing_ref, remarks)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        project_id,
                        material.get('row_no', ''),
                        material.get('material_name', ''),
                        material.get('spec', ''),
                        material.get('size', ''),
                        material.get('quantity', 0),
                        material.get('unit', ''),
                        material.get('construction_method', ''),
                        material.get('field_category', ''),
                        material.get('drawing_ref', ''),
                        material.get('remarks', '')
                    )
                )
            db.commit()
            stored_path = file_path
        elif file_ext == 'xls':
            file_type = 'material_list'
            # For .xls files, try to parse as Excel (may need xlrd)
            materials = parse_material_list_excel(file_path)
            for material in materials:
                cursor.execute(
                    '''INSERT INTO material_list
                       (project_id, row_no, material_name, spec, size, quantity, unit, construction_method, field_category, drawing_ref, remarks)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        project_id,
                        material.get('row_no', ''),
                        material.get('material_name', ''),
                        material.get('spec', ''),
                        material.get('size', ''),
                        material.get('quantity', 0),
                        material.get('unit', ''),
                        material.get('construction_method', ''),
                        material.get('field_category', ''),
                        material.get('drawing_ref', ''),
                        material.get('remarks', '')
                    )
                )
            db.commit()
            stored_path = file_path
        else:
            file_type = 'other'
            stored_path = file_path

        # Record in database
        cursor.execute(
            '''INSERT INTO project_files
               (project_id, file_type, original_name, stored_path, file_size, uploaded_by, uploaded_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (project_id, file_type, original_name, stored_path, file_size, current_user.id, datetime.utcnow())
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'UPLOAD_FILE',
            'project_file',
            cursor.lastrowid,
            'INFO',
            f'ファイルアップロード: {original_name} ({file_type})',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': 'ファイルがアップロードされました',
            'file_id': cursor.lastrowid
        }), 200

    except Exception as e:
        error_msg = str(e)
        # Parsing errors (Excel/CSV format issues) are user errors (400)
        # Server errors (disk, DB) are 500
        is_parse_error = any(kw in error_msg for kw in ['解析エラー', 'ヘッダー', 'File is not', 'not a zip'])
        status_code = 400 if is_parse_error else 500

        add_error_log(
            current_user.id,
            'UPLOAD_FILE_ERROR',
            error_msg,
            error_msg,
            request.url
        )
        return jsonify({'error': f'ファイルアップロードエラー: {error_msg}'}), status_code


@app.route('/projects/<int:project_id>/run-matching', methods=['POST'])
@login_required
def run_matching(project_id):
    """Execute matching engine"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        # Load master data
        load_master_data()

        # Run matching
        results = run_project_matching(project_id, current_user.id)

        # results is a dict, not a list, so get the count directly
        match_count = len(results) if isinstance(results, dict) else 0

        add_audit_log(
            current_user.id,
            'RUN_MATCHING',
            'project',
            project_id,
            'INFO',
            f'マッチング実行完了: {match_count}件',
            get_user_ip()
        )

        # Update project status
        cursor.execute(
            'UPDATE projects SET status = ?, updated_at = ? WHERE id = ?',
            ('matched', datetime.utcnow(), project_id)
        )
        db.commit()

        return jsonify({
            'success': True,
            'message': 'マッチング完了しました',
            'match_count': match_count
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'MATCHING_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'マッチングエラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/estimates')
@login_required
def get_estimates(project_id):
    """Get estimate details as JSON"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        cursor.execute(
            '''SELECT id, row_no, field_category, material_name, spec, construction_method, unit,
                      quantity, composite_unit_price, amount, productivity, productivity_total,
                      source_pdf, source_page, match_type, confidence, match_reason, remarks,
                      is_manual_added, material_id, master_id
               FROM estimate_details
               WHERE project_id = ?
               ORDER BY row_no''',
            (project_id,)
        )

        estimates = []
        for row in cursor.fetchall():
            estimates.append({
                'id': row[0],
                'row_no': row[1],
                'field_category': row[2],
                'material_name': row[3],
                'spec': row[4],
                'construction_method': row[5],
                'unit': row[6],
                'quantity': row[7],
                'composite_unit_price': row[8],
                'amount': row[9],
                'productivity': row[10],
                'productivity_total': row[11],
                'source_pdf': row[12],
                'source_page': row[13],
                'match_type': row[14],
                'confidence': row[15],
                'match_reason': row[16],
                'remarks': row[17],
                'is_manual_added': row[18],
                'material_id': row[19],
                'master_id': row[20]
            })

        return jsonify({'estimates': estimates}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'GET_ESTIMATES_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'見積取得エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/estimates/<int:detail_id>/edit', methods=['POST'])
@login_required
def edit_estimate_detail(project_id, detail_id):
    """Edit estimate detail and record history"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        data = request.get_json()
        column_name = data.get('column')
        new_value = data.get('value')

        if not column_name:
            return jsonify({'error': 'カラム名が必須です'}), 400

        # Get old value
        cursor.execute(
            f'SELECT {column_name} FROM estimate_details WHERE id = ? AND project_id = ?',
            (detail_id, project_id)
        )
        result = cursor.fetchone()
        if not result:
            return jsonify({'error': '見積詳細が見つかりません'}), 404

        old_value = result[0]

        # Handle amount recalculation when quantity or composite_unit_price is edited
        if column_name == 'composite_unit_price' or column_name == 'quantity':
            # Get the current values
            cursor.execute(
                'SELECT quantity, composite_unit_price, productivity FROM estimate_details WHERE id = ? AND project_id = ?',
                (detail_id, project_id)
            )
            current_vals = cursor.fetchone()
            if current_vals:
                current_quantity = current_vals[0]
                current_unit_price = current_vals[1]
                current_productivity = current_vals[2]

                # Update the specified column
                if column_name == 'quantity':
                    new_quantity = float(new_value) if new_value else 0
                    new_amount = new_quantity * current_unit_price
                    new_productivity_total = new_quantity * current_productivity if current_productivity else 0
                else:  # composite_unit_price
                    new_unit_price = float(new_value) if new_value else 0
                    new_amount = current_quantity * new_unit_price
                    new_productivity_total = current_quantity * current_productivity if current_productivity else 0

                # Update all three fields
                cursor.execute(
                    '''UPDATE estimate_details
                       SET quantity = ?, composite_unit_price = ?, amount = ?, productivity_total = ?
                       WHERE id = ? AND project_id = ?''',
                    (
                        float(new_quantity) if column_name == 'quantity' else current_quantity,
                        float(new_unit_price) if column_name == 'composite_unit_price' else current_unit_price,
                        new_amount,
                        new_productivity_total,
                        detail_id,
                        project_id
                    )
                )
        elif column_name == 'productivity':
            # When productivity is edited, recalculate productivity_total
            cursor.execute(
                'SELECT quantity FROM estimate_details WHERE id = ? AND project_id = ?',
                (detail_id, project_id)
            )
            qty_result = cursor.fetchone()
            if qty_result:
                quantity = qty_result[0]
                new_productivity = float(new_value) if new_value else 0
                new_productivity_total = quantity * new_productivity

                cursor.execute(
                    '''UPDATE estimate_details
                       SET productivity = ?, productivity_total = ?
                       WHERE id = ? AND project_id = ?''',
                    (new_productivity, new_productivity_total, detail_id, project_id)
                )
            else:
                cursor.execute(
                    f'UPDATE estimate_details SET {column_name} = ? WHERE id = ? AND project_id = ?',
                    (new_value, detail_id, project_id)
                )
        else:
            # Standard update
            cursor.execute(
                f'UPDATE estimate_details SET {column_name} = ? WHERE id = ? AND project_id = ?',
                (new_value, detail_id, project_id)
            )

        db.commit()

        # Record edit history
        cursor.execute(
            '''INSERT INTO edit_history
               (project_id, detail_id, column_name, old_value, new_value, edited_by, edited_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            (project_id, detail_id, column_name, str(old_value), str(new_value), current_user.id, datetime.utcnow())
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'EDIT_ESTIMATE',
            'estimate_detail',
            detail_id,
            'INFO',
            f'見積詳細編集: {column_name}',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': '更新しました'
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'EDIT_ESTIMATE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'編集エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/estimates/add-row', methods=['POST'])
@login_required
def add_estimate_row(project_id):
    """Add manual estimate row"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        data = request.get_json()

        # Get next row number
        cursor.execute(
            'SELECT MAX(row_no) FROM estimate_details WHERE project_id = ?',
            (project_id,)
        )
        result = cursor.fetchone()
        next_row_no = (result[0] or 0) + 1

        cursor.execute(
            '''INSERT INTO estimate_details
               (project_id, row_no, field_category, material_name, spec, construction_method, unit,
                quantity, composite_unit_price, amount, is_manual_added, edited_by)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
            (
                project_id,
                next_row_no,
                data.get('field_category', ''),
                data.get('material_name', ''),
                data.get('spec', ''),
                data.get('construction_method', ''),
                data.get('unit', ''),
                data.get('quantity', 0),
                data.get('composite_unit_price', 0),
                data.get('amount', 0),
                True,
                current_user.id
            )
        )
        db.commit()

        new_detail_id = cursor.lastrowid

        add_audit_log(
            current_user.id,
            'ADD_ESTIMATE_ROW',
            'estimate_detail',
            new_detail_id,
            'INFO',
            f'見積行追加（手動）',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': '行を追加しました',
            'detail_id': new_detail_id,
            'row_no': next_row_no
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'ADD_ROW_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'行追加エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/export-excel')
@login_required
def export_excel(project_id):
    """Export project to Excel"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by, name FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        project_name = project[1]

        # Create workbook
        wb = Workbook()

        # Sheet 1: 見積明細
        ws1 = wb.active
        ws1.title = '見積明細'

        headers_1 = [
            '行番号', '分野', '名称', '規格', '施工条件', '単位',
            '数量', '単価', '金額', '生産性', '生産性合計',
            'ソースPDF', 'ページ', 'マッチ種別', '信頼度', 'マッチ理由', '備考'
        ]
        ws1.append(headers_1)

        # Format header
        header_fill = PatternFill(start_color='0070C0', end_color='0070C0', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font

        cursor.execute(
            '''SELECT row_no, field_category, material_name, spec, construction_method, unit,
                      quantity, composite_unit_price, amount, productivity, productivity_total,
                      source_pdf, source_page, match_type, confidence, match_reason, remarks
               FROM estimate_details
               WHERE project_id = ?
               ORDER BY row_no''',
            (project_id,)
        )

        for row_data in cursor.fetchall():
            ws1.append(list(row_data))

        # Auto-filter
        ws1.auto_filter.ref = f'A1:{get_column_letter(len(headers_1))}1'

        # Sheet 2: 照合結果
        ws2 = wb.create_sheet('照合結果')

        headers_2 = [
            'プロジェクトID', '素材ID', '候補順位', 'マスターID', 'マッチ種別',
            '信頼度', '理由', '採用済', 'マスター名', 'マスター規格', 'マスター施工条件',
            'マスター単位', '単価', '生産性', 'ソースページ', '分野'
        ]
        ws2.append(headers_2)

        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font

        cursor.execute(
            '''SELECT project_id, material_id, candidate_rank, master_id, match_type,
                      confidence, reason, is_adopted, master_name, master_spec, master_method,
                      master_unit, composite_unit_price, removal_productivity, source_page, field_category
               FROM match_results
               WHERE project_id = ?''',
            (project_id,)
        )

        for row_data in cursor.fetchall():
            ws2.append(list(row_data))

        ws2.auto_filter.ref = f'A1:{get_column_letter(len(headers_2))}1'

        # Sheet 3: 修正履歴
        ws3 = wb.create_sheet('修正履歴')

        headers_3 = [
            'プロジェクトID', '見積詳細ID', 'カラム名', '旧値', '新値', '編集者ID', '編集日時'
        ]
        ws3.append(headers_3)

        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font

        cursor.execute(
            '''SELECT project_id, detail_id, column_name, old_value, new_value, edited_by, edited_at
               FROM edit_history
               WHERE project_id = ?
               ORDER BY edited_at DESC''',
            (project_id,)
        )

        for row_data in cursor.fetchall():
            ws3.append(list(row_data))

        ws3.auto_filter.ref = f'A1:{get_column_letter(len(headers_3))}1'

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        add_audit_log(
            current_user.id,
            'EXPORT_EXCEL',
            'project',
            project_id,
            'INFO',
            'Excel エクスポート',
            get_user_ip()
        )

        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'{project_name}_見積.xlsx'
        )

    except Exception as e:
        add_error_log(
            current_user.id,
            'EXPORT_EXCEL_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='Excel エクスポートエラーが発生しました'), 500


@app.route('/projects/<int:project_id>/match-candidates/<int:material_id>')
@login_required
def get_match_candidates(project_id, material_id):
    """Get match candidates for material as JSON"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        cursor.execute(
            '''SELECT id, candidate_rank, master_id, match_type, confidence, reason, is_adopted,
                      master_name, master_spec, master_method, master_unit, composite_unit_price,
                      removal_productivity, source_page, field_category
               FROM match_results
               WHERE project_id = ? AND material_id = ?
               ORDER BY candidate_rank''',
            (project_id, material_id)
        )

        candidates = []
        for row in cursor.fetchall():
            candidates.append({
                'id': row[0],
                'candidate_rank': row[1],
                'master_id': row[2],
                'match_type': row[3],
                'confidence': row[4],
                'reason': row[5],
                'is_adopted': row[6],
                'master_name': row[7],
                'master_spec': row[8],
                'master_method': row[9],
                'master_unit': row[10],
                'composite_unit_price': row[11],
                'removal_productivity': row[12],
                'source_page': row[13],
                'field_category': row[14]
            })

        return jsonify({'candidates': candidates}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'GET_CANDIDATES_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'候補取得エラー: {str(e)}'}), 500


# ==================== PROJECT ESTIMATE SETTINGS ROUTES ====================

@app.route('/projects/<int:project_id>/estimate-settings', methods=['GET', 'POST'])
@login_required
def estimate_settings(project_id):
    """View/edit project-specific estimate settings"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by, name FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        if request.method == 'POST':
            # Update or create project settings
            data = request.get_json() or request.form

            cursor.execute('SELECT id FROM project_estimate_settings WHERE project_id = ?', (project_id,))
            existing = cursor.fetchone()

            settings_data = {
                'company_name': data.get('company_name', ''),
                'company_address': data.get('company_address', ''),
                'company_tel': data.get('company_tel', ''),
                'company_fax': data.get('company_fax', ''),
                'labor_unit_price': float(data.get('labor_unit_price', 25000)) if data.get('labor_unit_price') else 25000,
                'estimate_title': data.get('estimate_title', ''),
                'estimate_conditions': data.get('estimate_conditions', ''),
                'updated_at': datetime.utcnow()
            }

            if existing:
                cursor.execute(
                    '''UPDATE project_estimate_settings
                       SET company_name = ?, company_address = ?, company_tel = ?, company_fax = ?,
                           labor_unit_price = ?, estimate_title = ?, estimate_conditions = ?, updated_at = ?
                       WHERE project_id = ?''',
                    (
                        settings_data['company_name'],
                        settings_data['company_address'],
                        settings_data['company_tel'],
                        settings_data['company_fax'],
                        settings_data['labor_unit_price'],
                        settings_data['estimate_title'],
                        settings_data['estimate_conditions'],
                        settings_data['updated_at'],
                        project_id
                    )
                )
            else:
                cursor.execute(
                    '''INSERT INTO project_estimate_settings
                       (project_id, company_name, company_address, company_tel, company_fax,
                        labor_unit_price, estimate_title, estimate_conditions, created_at, updated_at)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                    (
                        project_id,
                        settings_data['company_name'],
                        settings_data['company_address'],
                        settings_data['company_tel'],
                        settings_data['company_fax'],
                        settings_data['labor_unit_price'],
                        settings_data['estimate_title'],
                        settings_data['estimate_conditions'],
                        datetime.utcnow(),
                        settings_data['updated_at']
                    )
                )

            db.commit()

            add_audit_log(
                current_user.id,
                'UPDATE_PROJECT_SETTINGS',
                'project_settings',
                project_id,
                'INFO',
                f'案件見積設定更新: プロジェクト {project_id}',
                get_user_ip()
            )

            return jsonify({'success': True, 'message': '設定を保存しました'}), 200

        # GET: Return project settings (or empty if not set)
        cursor.execute(
            '''SELECT company_name, company_address, company_tel, company_fax,
                      labor_unit_price, estimate_title, estimate_conditions
               FROM project_estimate_settings
               WHERE project_id = ?''',
            (project_id,)
        )
        settings = cursor.fetchone()

        if settings:
            return jsonify({
                'company_name': settings[0],
                'company_address': settings[1],
                'company_tel': settings[2],
                'company_fax': settings[3],
                'labor_unit_price': settings[4],
                'estimate_title': settings[5],
                'estimate_conditions': settings[6]
            }), 200
        else:
            # Return empty settings
            return jsonify({
                'company_name': '',
                'company_address': '',
                'company_tel': '',
                'company_fax': '',
                'labor_unit_price': 25000,
                'estimate_title': '',
                'estimate_conditions': ''
            }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'ESTIMATE_SETTINGS_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'設定の取得に失敗しました: {str(e)}'}), 500


# ==================== ESTIMATE BUILDER ROUTES ====================

@app.route('/projects/<int:project_id>/estimate-builder')
@login_required
def estimate_builder(project_id):
    """Estimate builder page (見積作成画面)"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by, name FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return render_template('error.html', error_code=403, error_message='アクセス権限がありません'), 403

        # Get estimate details for the project
        cursor.execute(
            '''SELECT id, row_no, field_category, material_name, spec, construction_method, unit,
                      quantity, composite_unit_price, amount, productivity, productivity_total
               FROM estimate_details
               WHERE project_id = ?
               ORDER BY row_no''',
            (project_id,)
        )
        details = cursor.fetchall()

        # Get all settings (global defaults)
        settings = {}
        for row in cursor.execute("SELECT setting_key, setting_value FROM estimate_settings").fetchall():
            settings[row[0]] = row[1]

        # Check for project-specific settings and override defaults if they exist
        cursor.execute(
            'SELECT labor_unit_price FROM project_estimate_settings WHERE project_id = ?',
            (project_id,)
        )
        project_settings = cursor.fetchone()
        if project_settings and project_settings[0]:
            labor_unit_price = float(project_settings[0])
        else:
            labor_unit_price = float(settings.get('labor_unit_price', '25000'))

        # Convert to list of dicts
        details_list = []
        for row in details:
            details_list.append({
                'id': row[0],
                'row_no': row[1],
                'field_category': row[2],
                'material_name': row[3],
                'spec': row[4],
                'construction_method': row[5],
                'unit': row[6],
                'quantity': row[7],
                'composite_unit_price': row[8],
                'amount': row[9],
                'productivity': row[10],
                'productivity_total': row[11]
            })

        add_audit_log(
            current_user.id,
            'VIEW_ESTIMATE_BUILDER',
            'project',
            project_id,
            'INFO',
            '見積作成画面表示',
            get_user_ip()
        )

        # Build project dict for template
        project_dict = {
            'id': project_id,
            'name': project[1],
            'client_name': '',
        }
        # Get full project info
        cursor.execute('SELECT client_name, description FROM projects WHERE id = ?', (project_id,))
        pinfo = cursor.fetchone()
        if pinfo:
            project_dict['client_name'] = pinfo[0] or ''
            project_dict['description'] = pinfo[1] or ''

        return render_template(
            'estimate_builder.html',
            project=project_dict,
            details=details_list,
            settings=settings,
            labor_unit_price=labor_unit_price,
            today=datetime.now().strftime('%Y-%m-%d')
        )

    except Exception as e:
        add_error_log(
            current_user.id,
            'ESTIMATE_BUILDER_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='見積作成画面の読み込みに失敗しました'), 500


@app.route('/projects/<int:project_id>/estimate-builder/save', methods=['POST'])
@login_required
def save_estimate(project_id):
    """Save estimate data from estimate builder"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス稩限がありません'}), 403

        data = request.get_json()
        sections = data.get('sections', [])

        # Process and save sections/rows
        for section in sections:
            rows = section.get('rows', [])
            for row in rows:
                detail_id = row.get('id')
                if detail_id:
                    # Update existing row
                    cursor.execute(
                        '''UPDATE estimate_details
                           SET quantity = ?, composite_unit_price = ?, amount = ?,
                               productivity = ?, productivity_total = ?, remarks = ?
                           WHERE id = ? AND project_id = ?''',
                        (
                            float(row.get('quantity', 0)),
                            float(row.get('composite_unit_price', 0)),
                            float(row.get('amount', 0)),
                            float(row.get('productivity', 0)),
                            float(row.get('productivity_total', 0)),
                            row.get('remarks', ''),
                            detail_id,
                            project_id
                        )
                    )
                else:
                    # Insert new row
                    cursor.execute(
                        '''INSERT INTO estimate_details
                           (project_id, row_no, field_category, material_name, spec,
                            construction_method, unit, quantity, composite_unit_price,
                            amount, productivity, productivity_total, remarks, is_manual_added)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        (
                            project_id,
                            row.get('row_no', 0),
                            row.get('field_category', ''),
                            row.get('material_name', ''),
                            row.get('spec', ''),
                            row.get('construction_method', ''),
                            row.get('unit', ''),
                            float(row.get('quantity', 0)),
                            float(row.get('composite_unit_price', 0)),
                            float(row.get('amount', 0)),
                            float(row.get('productivity', 0)),
                            float(row.get('productivity_total', 0)),
                            row.get('remarks', ''),
                            True
                        )
                    )

        db.commit()

        add_audit_log(
            current_user.id,
            'SAVE_ESTIMATE',
            'project',
            project_id,
            'INFO',
            '見積データ保存',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': '見積データを保存しました'
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'SAVE_ESTIMATE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'保存エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/estimate-builder/export-pdf', methods=['POST'])
@login_required
def export_estimate_pdf(project_id):
    """Export estimate as Excel (PDF generation is complex, so use Excel)"""
    try:
        # For now, just redirect to the existing Excel export endpoint
        return redirect(url_for('export_excel', project_id=project_id))
    except Exception as e:
        add_error_log(
            current_user.id,
            'EXPORT_PDF_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'エクスポートエラー: {str(e)}'}), 500


# ==================== SHARED FILES ROUTES ====================

@app.route('/projects/<int:project_id>/files')
@login_required
def project_files(project_id):
    """Get project files as JSON"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        cursor.execute(
            '''SELECT id, file_type, original_name, file_size, uploaded_at, uploaded_by
               FROM project_files
               WHERE project_id = ?
               ORDER BY uploaded_at DESC''',
            (project_id,)
        )

        files = []
        for row in cursor.fetchall():
            files.append({
                'id': row[0],
                'file_type': row[1],
                'original_name': row[2],
                'file_size': row[3],
                'uploaded_at': row[4],
                'uploaded_by': row[5]
            })

        return jsonify({'files': files}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'GET_FILES_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'ファイル取得エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/files/download/<int:file_id>')
@login_required
def download_file(project_id, file_id):
    """Download a specific file"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return render_template('error.html', error_code=403, error_message='アクセス権限がありません'), 403

        # Get file info
        cursor.execute(
            'SELECT stored_path, original_name FROM project_files WHERE id = ? AND project_id = ?',
            (file_id, project_id)
        )
        file_info = cursor.fetchone()
        if not file_info:
            return render_template('error.html', error_code=404, error_message='ファイルが見つかりません'), 404

        stored_path, original_name = file_info

        if not os.path.exists(stored_path):
            return render_template('error.html', error_code=404, error_message='ファイルが存在しません'), 404

        add_audit_log(
            current_user.id,
            'DOWNLOAD_FILE',
            'project_file',
            file_id,
            'INFO',
            f'ファイルダウンロード: {original_name}',
            get_user_ip()
        )

        return send_file(
            stored_path,
            as_attachment=True,
            download_name=original_name
        )

    except Exception as e:
        add_error_log(
            current_user.id,
            'DOWNLOAD_FILE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='ファイルダウンロードエラーが発生しました'), 500


# ==================== LEARNING DICTIONARY ROUTES ====================

@app.route('/projects/<int:project_id>/add-to-learning', methods=['POST'])
@login_required
def add_to_learning(project_id):
    """Add corrected match to learning dictionary"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        data = request.get_json()

        cursor.execute(
            '''INSERT INTO learning_dictionary
               (input_name, canonical_name, input_spec, canonical_spec,
                input_method, canonical_method, status, source_project_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
            (
                data.get('input_name', ''),
                data.get('canonical_name', ''),
                data.get('input_spec', ''),
                data.get('canonical_spec', ''),
                data.get('input_method', ''),
                data.get('canonical_method', ''),
                'candidate',
                project_id
            )
        )
        db.commit()

        new_entry_id = cursor.lastrowid

        add_audit_log(
            current_user.id,
            'ADD_LEARNING_ENTRY',
            'project',
            project_id,
            'INFO',
            f'学習辞書エントリ追加',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': '学習辞書に追加しました',
            'entry_id': new_entry_id
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'ADD_LEARNING_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'追加エラー: {str(e)}'}), 500


# ==================== ADMIN ROUTES ====================

@app.route('/admin/users')
@admin_required
def admin_users():
    """User management page"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            '''SELECT id, email, full_name, role, is_active, created_at, approved_at, approved_by, last_login_at
               FROM users
               ORDER BY created_at DESC''',
            ()
        )

        users = []
        for row in cursor.fetchall():
            users.append({
                'id': row[0],
                'email': row[1],
                'full_name': row[2],
                'role': row[3],
                'is_active': row[4],
                'created_at': row[5],
                'approved_at': row[6],
                'approved_by': row[7],
                'last_login_at': row[8]
            })

        add_audit_log(
            current_user.id,
            'VIEW_USER_MANAGEMENT',
            'admin',
            None,
            'INFO',
            'ユーザー管理ページ表示',
            get_user_ip()
        )

        return render_template('admin_users.html', users=users)

    except Exception as e:
        add_error_log(
            current_user.id,
            'ADMIN_USERS_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='ユーザー管理読み込みエラーが発生しました'), 500


@app.route('/admin/users/<int:user_id>/approve', methods=['POST'])
@admin_required
def approve_user(user_id):
    """Approve pending user and send notification email"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Get user info before update
        cursor.execute('SELECT email, full_name FROM users WHERE id = ? AND role = ?', (user_id, 'pending'))
        user_row = cursor.fetchone()
        if not user_row:
            return jsonify({'error': 'ユーザーが見つからないか、既に承認済みです'}), 404

        user_email = user_row[0]
        user_name = user_row[1]

        cursor.execute(
            'UPDATE users SET role = ?, approved_at = ?, approved_by = ? WHERE id = ? AND role = ?',
            ('user', datetime.utcnow(), current_user.id, user_id, 'pending')
        )
        db.commit()

        # Send approval notification email
        email_sent = send_approval_email(user_email, user_name, triggered_by=current_user.id, related_user_id=user_id)

        add_audit_log(
            current_user.id,
            'APPROVE_USER',
            'user',
            user_id,
            'INFO',
            f'ユーザー承認 (メール通知: {"成功" if email_sent else "未送信/失敗"})',
            get_user_ip()
        )

        message = 'ユーザーを承認しました'
        if email_sent:
            message += f'。{user_email} に通知メールを送信しました'
        else:
            message += '。（メール通知は未設定のため送信されていません）'

        return jsonify({'success': True, 'message': message, 'email_sent': email_sent}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'APPROVE_USER_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'承認エラー: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/reject', methods=['POST'])
@admin_required
def reject_user(user_id):
    """Reject pending user and send notification email"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Get user info before deletion
        cursor.execute('SELECT email, full_name FROM users WHERE id = ? AND role = ?', (user_id, 'pending'))
        user_row = cursor.fetchone()
        if not user_row:
            return jsonify({'error': 'ユーザーが見つからないか、既に処理済みです'}), 404

        user_email = user_row[0]
        user_name = user_row[1]

        # Get rejection reason from request body
        data = request.get_json(silent=True) or {}
        reason = data.get('reason', '')

        # Send rejection notification email BEFORE deleting the user
        email_sent = send_rejection_email(user_email, user_name, reason, triggered_by=current_user.id, related_user_id=user_id)

        cursor.execute(
            'DELETE FROM users WHERE id = ? AND role = ?',
            (user_id, 'pending')
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'REJECT_USER',
            'user',
            user_id,
            'INFO',
            f'ユーザー却下: {user_name} ({user_email}) 理由: {reason or "なし"} メール: {"送信" if email_sent else "未送信"}',
            get_user_ip()
        )

        message = f'{user_name} さんを却下しました'
        if email_sent:
            message += f'。{user_email} に通知メールを送信しました'
        else:
            message += '。（メール通知は未設定のため送信されていません）'

        return jsonify({'success': True, 'message': message, 'email_sent': email_sent}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'REJECT_USER_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'却下エラー: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/toggle-active', methods=['POST'])
@admin_required
def toggle_user_active(user_id):
    """Toggle user active status"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute('SELECT is_active FROM users WHERE id = ?', (user_id,))
        result = cursor.fetchone()
        if not result:
            return jsonify({'error': 'ユーザーが見つかりません'}), 404

        new_status = not result[0]
        cursor.execute(
            'UPDATE users SET is_active = ? WHERE id = ?',
            (new_status, user_id)
        )
        db.commit()

        status_text = 'アクティブ化' if new_status else '無効化'
        add_audit_log(
            current_user.id,
            'TOGGLE_USER_ACTIVE',
            'user',
            user_id,
            'INFO',
            f'ユーザー{status_text}',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': f'ユーザーを{status_text}しました', 'is_active': new_status}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'TOGGLE_ACTIVE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'状態変更エラー: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/change-role', methods=['POST'])
@admin_required
def change_user_role(user_id):
    """Change user role"""
    try:
        db = get_db()
        cursor = db.cursor()

        data = request.get_json()
        new_role = data.get('role')

        if new_role not in ['user', 'admin']:
            return jsonify({'error': '無効なロールです'}), 400

        cursor.execute(
            'UPDATE users SET role = ? WHERE id = ?',
            (new_role, user_id)
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'CHANGE_USER_ROLE',
            'user',
            user_id,
            'INFO',
            f'ロール変更: {new_role}',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'ロールを変更しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'CHANGE_ROLE_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'ロール変更エラー: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/reset-password', methods=['POST'])
@admin_required
def reset_user_password(user_id):
    """Reset user password"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check that user exists and is not the current admin
        if user_id == current_user.id:
            return jsonify({'error': '自分自身のパスワードはこの方法ではリセットできません'}), 400

        cursor.execute('SELECT email FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'error': 'ユーザーが見つかりません'}), 404

        # Generate temporary password
        import string
        import secrets
        temp_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))

        # Hash and update password
        pw_hash = bcrypt.hashpw(temp_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        cursor.execute(
            'UPDATE users SET password_hash = ? WHERE id = ?',
            (pw_hash, user_id)
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'RESET_USER_PASSWORD',
            'user',
            user_id,
            'INFO',
            'ユーザーパスワードリセット',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': 'パスワードをリセットしました',
            'temp_password': temp_password
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'RESET_PASSWORD_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'パスワードリセットエラー: {str(e)}'}), 500


@app.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@admin_required
def delete_user(user_id):
    """Delete user"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check that user is not the current admin
        if user_id == current_user.id:
            return jsonify({'error': '自分自身を削除することはできません'}), 400

        cursor.execute('SELECT email FROM users WHERE id = ?', (user_id,))
        user = cursor.fetchone()
        if not user:
            return jsonify({'error': 'ユーザーが見つかりません'}), 404

        # Delete user
        cursor.execute('DELETE FROM users WHERE id = ?', (user_id,))
        db.commit()

        add_audit_log(
            current_user.id,
            'DELETE_USER',
            'user',
            user_id,
            'INFO',
            f'ユーザー削除: {user[0]}',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'ユーザーを削除しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'DELETE_USER_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'ユーザー削除エラー: {str(e)}'}), 500


@app.route('/admin/email-log')
@admin_required
def admin_email_log():
    """Email send log viewer for admin"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            '''SELECT e.id, e.to_email, e.to_name, e.subject, e.email_type, e.status,
                      e.error_message, e.created_at,
                      COALESCE(u.full_name, 'システム') as triggered_by_name
               FROM email_log e
               LEFT JOIN users u ON e.triggered_by = u.id
               ORDER BY e.created_at DESC
               LIMIT 100''',
            ()
        )

        logs = []
        for row in cursor.fetchall():
            logs.append({
                'id': row[0],
                'to_email': row[1],
                'to_name': row[2],
                'subject': row[3],
                'email_type': row[4],
                'status': row[5],
                'error_message': row[6],
                'created_at': row[7],
                'triggered_by_name': row[8],
            })

        # Email type labels
        type_labels = {
            'registration_admin_notify': '登録通知（管理者宛）',
            'registration_user_confirm': '登録確認（ユーザー宛）',
            'approval_notify': '承認通知',
            'rejection_notify': '却下通知',
            'general': 'その他',
        }

        # Status labels
        status_labels = {
            'sent': '送信成功',
            'failed': '送信失敗',
            'skipped': 'スキップ',
            'pending': '保留中',
        }

        add_audit_log(
            current_user.id,
            'VIEW_EMAIL_LOG',
            'admin',
            None,
            'INFO',
            'メール送信ログ表示',
            get_user_ip()
        )

        return render_template('admin_email_log.html', logs=logs, type_labels=type_labels, status_labels=status_labels)

    except Exception as e:
        add_error_log(
            current_user.id,
            'EMAIL_LOG_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='メール送信ログ読み込みエラーが発生しました'), 500


@app.route('/admin/audit-log')
@admin_required
def admin_audit_log():
    """Audit log viewer"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Get filters
        user_id = request.args.get('user_id', '')
        action = request.args.get('action', '')
        level = request.args.get('level', '')

        query = '''SELECT a.id, a.user_id, a.action, a.entity_type, a.entity_id,
                   a.level, a.details, a.created_at, COALESCE(u.full_name,'システム') as user_name
                   FROM audit_log a LEFT JOIN users u ON a.user_id=u.id'''
        params = []

        conditions = []
        if user_id:
            conditions.append('a.user_id = ?')
            params.append(user_id)
        if action:
            conditions.append('a.action LIKE ?')
            params.append(f'%{action}%')
        if level:
            conditions.append('a.level = ?')
            params.append(level)

        if conditions:
            query += ' WHERE ' + ' AND '.join(conditions)

        # Pagination
        page = int(request.args.get('page', 1))
        per_page = 50
        count_query = query.replace(
            'SELECT a.id, a.user_id, a.action, a.entity_type, a.entity_id,\n                   a.level, a.details, a.created_at, COALESCE(u.full_name,\'システム\') as user_name\n                   FROM audit_log a LEFT JOIN users u ON a.user_id=u.id',
            'SELECT COUNT(*) FROM audit_log a LEFT JOIN users u ON a.user_id=u.id'
        )
        cursor.execute(count_query, params)
        total = cursor.fetchone()[0]
        total_pages = max(1, (total + per_page - 1) // per_page)

        query += f' ORDER BY a.created_at DESC LIMIT {per_page} OFFSET {(page-1)*per_page}'

        cursor.execute(query, params)

        logs = []
        for row in cursor.fetchall():
            logs.append({
                'id': row[0],
                'user_id': row[1],
                'action': row[2],
                'target': f"{row[3] or ''} {row[4] or ''}".strip(),
                'level': row[5],
                'details': row[6],
                'timestamp': row[7],
                'user_name': row[8]
            })

        add_audit_log(
            current_user.id,
            'VIEW_AUDIT_LOG',
            'admin',
            None,
            'INFO',
            '監査ログ表示',
            get_user_ip()
        )

        return render_template('admin_audit.html', logs=logs, page=page, total_pages=total_pages)

    except Exception as e:
        add_error_log(
            current_user.id,
            'AUDIT_LOG_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='監査ログ読み込みエラーが発生しました'), 500


@app.route('/admin/error-log')
@admin_required
def admin_error_log():
    """Error log viewer"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            '''SELECT e.id, e.user_id, e.error_type, e.error_message, e.traceback, e.url, e.created_at,
                      COALESCE(u.full_name,'システム') as user_name
               FROM error_log e LEFT JOIN users u ON e.user_id=u.id
               ORDER BY e.created_at DESC
               LIMIT 500'''
        )

        logs = []
        for row in cursor.fetchall():
            logs.append({
                'id': row[0],
                'user_id': row[1],
                'error_type': row[2],
                'message': row[3],
                'traceback': row[4],
                'url': row[5],
                'timestamp': row[6],
                'user_name': row[7]
            })

        add_audit_log(
            current_user.id,
            'VIEW_ERROR_LOG',
            'admin',
            None,
            'INFO',
            'エラーログ表示',
            get_user_ip()
        )

        return render_template('admin_error.html', logs=logs)

    except Exception as e:
        add_error_log(
            current_user.id,
            'ERROR_LOG_VIEW_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='エラーログ読み込みエラーが発生しました'), 500


@app.route('/admin/learning')
@admin_required
def admin_learning():
    """Learning dictionary management"""
    try:
        db = get_db()
        cursor = db.cursor()

        status_filter = request.args.get('status', 'candidate')

        cursor.execute(
            '''SELECT id, input_name, canonical_name, input_spec, canonical_spec, input_method,
                      canonical_method, confidence, status, confirmed_by, confirmed_at, source_project_id
               FROM learning_dictionary
               WHERE status = ?
               ORDER BY confidence DESC''',
            (status_filter,)
        )

        entries = []
        for row in cursor.fetchall():
            entries.append({
                'id': row[0],
                'input_name': row[1],
                'canonical_name': row[2],
                'input_spec': row[3],
                'canonical_spec': row[4],
                'input_method': row[5],
                'canonical_method': row[6],
                'confidence': row[7],
                'status': row[8],
                'confirmed_by': row[9],
                'confirmed_at': row[10],
                'source_project_id': row[11]
            })

        add_audit_log(
            current_user.id,
            'VIEW_LEARNING_DICTIONARY',
            'admin',
            None,
            'INFO',
            f'学習辞書表示 ({status_filter})',
            get_user_ip()
        )

        return render_template('admin_learning.html', entries=entries, status_filter=status_filter)

    except Exception as e:
        add_error_log(
            current_user.id,
            'LEARNING_DICT_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='学習辞書読み込みエラーが発生しました'), 500


@app.route('/admin/learning/<int:entry_id>/confirm', methods=['POST'])
@admin_required
def confirm_learning_entry(entry_id):
    """Confirm learning dictionary entry"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            '''UPDATE learning_dictionary
               SET status = ?, confirmed_by = ?, confirmed_at = ?
               WHERE id = ?''',
            ('confirmed', current_user.id, datetime.utcnow(), entry_id)
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'CONFIRM_LEARNING_ENTRY',
            'learning_dictionary',
            entry_id,
            'INFO',
            '学習辞書エントリ確認',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'エントリを確認しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'CONFIRM_LEARNING_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'確認エラー: {str(e)}'}), 500


@app.route('/admin/learning/<int:entry_id>/reject', methods=['POST'])
@admin_required
def reject_learning_entry(entry_id):
    """Reject learning dictionary entry"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            '''UPDATE learning_dictionary
               SET status = ?, confirmed_by = ?, confirmed_at = ?
               WHERE id = ?''',
            ('rejected', current_user.id, datetime.utcnow(), entry_id)
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'REJECT_LEARNING_ENTRY',
            'learning_dictionary',
            entry_id,
            'INFO',
            '学習辞書エントリ拒否',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': 'エントリを拒否しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'REJECT_LEARNING_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'拒否エラー: {str(e)}'}), 500


@app.route('/admin/master')
@admin_required
def admin_master():
    """Master data management"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Get latest update log
        cursor.execute(
            '''SELECT id, action, source_file, records_added, records_updated, status, updated_by, updated_at
               FROM master_update_log
               ORDER BY updated_at DESC
               LIMIT 50''',
            ()
        )

        logs = []
        for row in cursor.fetchall():
            logs.append({
                'id': row[0],
                'action': row[1],
                'source_file': row[2],
                'records_added': row[3],
                'records_updated': row[4],
                'status': row[5],
                'updated_by': row[6],
                'updated_at': row[7]
            })

        # Get master data stats
        cursor.execute(
            '''SELECT COUNT(*), COUNT(DISTINCT category_no), COUNT(DISTINCT field_category)
               FROM estimate_master'''
        )
        stats = cursor.fetchone()

        add_audit_log(
            current_user.id,
            'VIEW_MASTER_DATA',
            'admin',
            None,
            'INFO',
            'マスターデータ管理ページ表示',
            get_user_ip()
        )

        return render_template(
            'admin_master.html',
            logs=logs,
            total_records=stats[0] if stats else 0,
            total_categories=stats[1] if stats else 0,
            total_fields=stats[2] if stats else 0
        )

    except Exception as e:
        add_error_log(
            current_user.id,
            'MASTER_DATA_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='マスターデータ読み込みエラーが発生しました'), 500


@app.route('/admin/master/upload', methods=['POST'])
@admin_required
def upload_master_data():
    """Upload master data (Excel, CSV, TSV, SHD, etc.)"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        # Determine file extension
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''

        # Check if file type is supported
        if file_ext not in ('xlsx', 'xls', 'csv', 'tsv', 'shd', 'txt', 'pdf'):
            return jsonify({'error': '.xlsx, .xls, .csv, .tsv, .shd, .txt, .pdf ファイルのみサポートしています'}), 400

        # Save temp file
        temp_filename = f"master_import_{uuid.uuid4().hex}.{file_ext}"
        temp_path = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        file.save(temp_path)

        db = get_db()
        cursor = db.cursor()

        records_added = 0
        records_updated = 0
        rows_data = []

        # Parse file based on type
        if file_ext in ('xlsx', 'xls'):
            from openpyxl import load_workbook
            wb = load_workbook(temp_path)
            ws = wb.active
            # Skip header row
            for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_idx == 1:
                    continue
                if all(cell is None for cell in row):
                    continue
                rows_data.append(row)
        elif file_ext in ('csv', 'tsv', 'txt'):
            # Parse CSV/TSV
            delimiter = '\t' if file_ext == 'tsv' else ','
            with open(temp_path, 'r', encoding='utf-8', errors='replace') as f:
                reader = csv.reader(f, delimiter=delimiter)
                for row_idx, row in enumerate(reader):
                    if row_idx == 0:  # Skip header
                        continue
                    if not any(row):  # Skip empty rows
                        continue
                    rows_data.append(row)
        elif file_ext == 'shd':
            # Parse SHD file
            with open(temp_path, 'rb') as f:
                content = f.read().decode('shift_jis', errors='replace')
            lines = content.split('\r')
            for row_idx, line in enumerate(lines):
                if row_idx == 0:  # Skip header
                    continue
                fields = line.split('\t')
                if not any(fields):
                    continue
                rows_data.append(fields)
        elif file_ext == 'pdf':
            # For PDF, extract text and skip (not ideal for master data)
            return jsonify({'error': 'PDFファイルはマスターデータアップロードに適していません'}), 400

        # Process rows
        for row_data in rows_data:
            try:
                # Prepare insert data based on expected format
                data = {
                    'source_page': row_data[0] if len(row_data) > 0 else '',
                    'category_no': row_data[1] if len(row_data) > 1 else '',
                    'field_category': row_data[2] if len(row_data) > 2 else '',
                    'material_name': row_data[3] if len(row_data) > 3 else '',
                    'spec_summary': row_data[4] if len(row_data) > 4 else '',
                    'remarks': row_data[5] if len(row_data) > 5 else '',
                    'construction_method': row_data[6] if len(row_data) > 6 else '',
                    'unit': row_data[7] if len(row_data) > 7 else '',
                    'material_unit_price': float(row_data[8]) if len(row_data) > 8 and row_data[8] else 0,
                    'material_cost': float(row_data[9]) if len(row_data) > 9 and row_data[9] else 0,
                    'labor_cost': float(row_data[10]) if len(row_data) > 10 and row_data[10] else 0,
                    'expense_cost': float(row_data[11]) if len(row_data) > 11 and row_data[11] else 0,
                    'composite_unit_price': float(row_data[12]) if len(row_data) > 12 and row_data[12] else 0,
                    'removal_productivity': float(row_data[13]) if len(row_data) > 13 and row_data[13] else 0,
                    'removal_cost': float(row_data[14]) if len(row_data) > 14 and row_data[14] else 0,
                    'master_version': 1,
                }

                # Check if exists
                cursor.execute(
                    'SELECT id FROM estimate_master WHERE category_no = ? AND material_name = ?',
                    (data['category_no'], data['material_name'])
                )
                if cursor.fetchone():
                    records_updated += 1
                else:
                    cursor.execute(
                        '''INSERT INTO estimate_master
                           (source_page, category_no, field_category, material_name, spec_summary, remarks,
                            construction_method, unit, material_unit_price, material_cost, labor_cost,
                            expense_cost, composite_unit_price, removal_productivity, removal_cost, master_version)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                        tuple(data.values())
                    )
                    records_added += 1

            except Exception as row_error:
                continue

        db.commit()
        os.remove(temp_path)

        # Record update log
        cursor.execute(
            '''INSERT INTO master_update_log
               (action, source_file, records_added, records_updated, status, updated_by, updated_at)
               VALUES (?, ?, ?, ?, ?, ?, ?)''',
            ('IMPORT', file.filename, records_added, records_updated, 'success', current_user.id, datetime.utcnow())
        )
        db.commit()

        add_audit_log(
            current_user.id,
            'UPLOAD_MASTER_DATA',
            'master_data',
            None,
            'INFO',
            f'マスターデータアップロード: {records_added} 追加, {records_updated} 更新',
            get_user_ip()
        )

        return jsonify({
            'success': True,
            'message': 'マスターデータをアップロードしました',
            'records_added': records_added,
            'records_updated': records_updated
        }), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'UPLOAD_MASTER_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'マスターデータアップロードエラー: {str(e)}'}), 500


@app.route('/admin/settings')
@admin_required
def admin_settings():
    """Estimate settings management"""
    try:
        db = get_db()
        cursor = db.cursor()

        cursor.execute(
            '''SELECT id, setting_key, setting_value, description
               FROM estimate_settings
               ORDER BY setting_key''',
            ()
        )

        settings = {}
        for row in cursor.fetchall():
            settings[row[1]] = {
                'id': row[0],
                'value': row[2],
                'description': row[3]
            }

        add_audit_log(
            current_user.id,
            'VIEW_SETTINGS',
            'admin',
            None,
            'INFO',
            '設定管理ページ表示',
            get_user_ip()
        )

        return render_template('admin_settings.html', settings=settings)

    except Exception as e:
        add_error_log(
            current_user.id,
            'SETTINGS_ERROR',
            str(e),
            str(e),
            request.url
        )
        return render_template('error.html', error_code=500, error_message='設定読み込みエラーが発生しました'), 500


@app.route('/admin/settings', methods=['POST'])
@admin_required
def update_settings():
    """Update estimate settings"""
    try:
        db = get_db()
        cursor = db.cursor()

        data = request.get_json()

        for setting_key, setting_value in data.items():
            cursor.execute(
                '''UPDATE estimate_settings
                   SET setting_value = ?, updated_by = ?, updated_at = ?
                   WHERE setting_key = ?''',
                (str(setting_value), current_user.id, datetime.utcnow(), setting_key)
            )

        db.commit()

        add_audit_log(
            current_user.id,
            'UPDATE_SETTINGS',
            'admin',
            None,
            'INFO',
            '設定更新',
            get_user_ip()
        )

        return jsonify({'success': True, 'message': '設定を更新しました'}), 200

    except Exception as e:
        add_error_log(
            current_user.id,
            'UPDATE_SETTINGS_ERROR',
            str(e),
            str(e),
            request.url
        )
        return jsonify({'error': f'設定更新エラー: {str(e)}'}), 500


# ==================== BLUEPRINT VIEWER ====================

@app.route('/projects/<int:project_id>/blueprint/<int:file_id>')
@login_required
def blueprint_viewer(project_id, file_id):
    """Blueprint viewer with PDF display and material picking"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project ownership
        cursor.execute('SELECT id, name, client_name, description, status FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project:
            return render_template('error.html', error_code=404, error_message='プロジェクトが見つかりません'), 404

        # Get file info
        cursor.execute(
            'SELECT id, file_type, original_name FROM project_files WHERE id = ? AND project_id = ?',
            (file_id, project_id)
        )
        file_row = cursor.fetchone()
        if not file_row:
            return render_template('error.html', error_code=404, error_message='ファイルが見つかりません'), 404

        # Get existing blueprint items
        cursor.execute(
            '''SELECT id, page_number, material_name, spec, quantity, unit,
                      construction_method, field_category, confidence, match_type, reason, is_adopted
               FROM blueprint_items
               WHERE project_id = ? AND file_id = ?
               ORDER BY id''',
            (project_id, file_id)
        )
        materials = [dict(r) for r in cursor.fetchall()]

        # Get master data count
        cursor.execute('SELECT COUNT(*) FROM estimate_master')
        master_count = cursor.fetchone()[0]

        return render_template(
            'blueprint_viewer.html',
            project={'id': project[0], 'name': project[1], 'client_name': project[2],
                     'description': project[3], 'status': project[4]},
            file={'id': file_row[0], 'file_type': file_row[1], 'original_name': file_row[2]},
            materials=materials,
            master_count=master_count,
            pdf_url=url_for('serve_file', project_id=project_id, file_id=file_id)
        )

    except Exception as e:
        add_error_log(current_user.id, 'BLUEPRINT_VIEWER_ERROR', str(e), str(e), request.url)
        return render_template('error.html', error_code=500, error_message='図面ビューア読み込みエラー'), 500


@app.route('/projects/<int:project_id>/file/<int:file_id>/serve')
@login_required
def serve_file(project_id, file_id):
    """Serve a project file for viewing in browser"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            'SELECT stored_path, original_name, file_type FROM project_files WHERE id = ? AND project_id = ?',
            (file_id, project_id)
        )
        file_row = cursor.fetchone()
        if not file_row:
            return jsonify({'error': 'ファイルが見つかりません'}), 404

        stored_path = file_row[0]
        original_name = file_row[1]
        ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''

        mime_types = {
            'pdf': 'application/pdf',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'xls': 'application/vnd.ms-excel',
            'csv': 'text/csv',
            'txt': 'text/plain',
        }
        mimetype = mime_types.get(ext, 'application/octet-stream')

        return send_file(stored_path, mimetype=mimetype, as_attachment=False, download_name=original_name)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/projects/<int:project_id>/ai-extract/<int:file_id>', methods=['POST'])
@login_required
def ai_extract(project_id, file_id):
    """AI extraction of electrical equipment from PDF blueprint"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Get file path
        cursor.execute(
            'SELECT stored_path, original_name FROM project_files WHERE id = ? AND project_id = ?',
            (file_id, project_id)
        )
        file_row = cursor.fetchone()
        if not file_row:
            return jsonify({'error': 'ファイルが見つかりません'}), 404

        stored_path = file_row[0]

        # Check if we already have cached text
        cursor.execute('SELECT page_number, text_content FROM pdf_page_text WHERE file_id = ? ORDER BY page_number', (file_id,))
        cached_pages = cursor.fetchall()

        if cached_pages:
            page_texts = {r[0]: r[1] for r in cached_pages}
        else:
            # Extract text from PDF
            page_texts = {}
            try:
                with pdfplumber.open(stored_path) as pdf:
                    for page_num, page in enumerate(pdf.pages, 1):
                        text = page.extract_text() or ""
                        page_texts[page_num] = text
                        # Cache it
                        cursor.execute(
                            'INSERT INTO pdf_page_text (file_id, page_number, text_content) VALUES (?,?,?)',
                            (file_id, page_num, text)
                        )
                db.commit()
            except Exception as pdf_err:
                return jsonify({'error': f'PDF解析エラー: {str(pdf_err)}'}), 500

        # AI extraction: parse electrical equipment from text
        extracted_items = extract_electrical_equipment(page_texts)

        # Match against master data
        master_data = load_master_data()
        if master_data:
            from matching_engine import build_indexes, match_single_material
            indexes = build_indexes(master_data)

            for item in extracted_items:
                candidates = match_single_material(item, master_data, indexes, max_candidates=3)
                if candidates and candidates[0]['confidence'] > 0:
                    best = candidates[0]
                    item['confidence'] = best['confidence']
                    item['match_type'] = best['match_type']
                    item['reason'] = best['reason']
                    item['master_id'] = best['master_id']
                    item['master_name'] = best['master_name']
                    item['candidates'] = candidates
                else:
                    item['confidence'] = 0.3
                    item['match_type'] = 'ai_detected'
                    item['reason'] = 'AI図面解析による検出（マスタ未照合）'
                    item['candidates'] = []

        return jsonify({
            'success': True,
            'items': extracted_items,
            'total_pages': len(page_texts),
            'message': f'{len(extracted_items)}件の電気設備を検出しました'
        }), 200

    except Exception as e:
        add_error_log(current_user.id, 'AI_EXTRACT_ERROR', str(e), str(e), request.url)
        return jsonify({'error': f'AI解析エラー: {str(e)}'}), 500


def extract_electrical_equipment(page_texts):
    """Extract electrical equipment items from PDF text using pattern matching"""
    import re

    # Electrical equipment patterns
    patterns = {
        'ケーブル': [
            r'(CV[VTFS]*)\s*[-]?\s*(\d+(?:\.\d+)?)\s*(?:sq|mm|㎟)?(?:\s*[-x×]\s*(\d+)C?)?',
            r'(VVF|VVR|IV|HIV|EM[-\s]?(?:CE|EEF|IC|IE))\s*[-]?\s*(\d+(?:\.\d+)?)\s*(?:sq|mm)?(?:\s*[-x×]\s*(\d+)C?)?',
            r'(VCTF|CVVS?|CPEV|AE)\s*[-]?\s*(\d+(?:\.\d+)?)\s*[-x×]\s*(\d+)(?:C|P)?',
        ],
        '電線管': [
            r'(E|G|C|PF|CD|HIVE|FEP|VE)\s*(\d+)',
            r'(薄鋼電線管|厚鋼電線管|ねじなし電線管|合成樹脂管|PF管|CD管|FEP管|硬質ビニル管)\s*[-]?\s*(\d+)',
        ],
        'ケーブルラック': [
            r'(ケーブルラック|cable\s*rack)\s*[-]?\s*(\d+)\s*[x×]\s*(\d+)',
            r'(はしご形|トレー形|メッシュ形).*?ラック\s*[-]?\s*(\d+)',
        ],
        '配電盤・分電盤': [
            r'(配電盤|分電盤|動力盤|電灯盤|制御盤|MCC|MCCB|P[-\s]?\d+|L[-\s]?\d+)',
        ],
        '照明器具': [
            r'(LED[照明灯具]*|蛍光灯|ダウンライト|シーリング|スポットライト|非常灯|誘導灯)',
            r'(\d+W|ワット)\s*(LED|蛍光)',
        ],
        'コンセント・スイッチ': [
            r'(コンセント|アウトレット|タンブラスイッチ|TS|WN\d+)',
            r'(1P|2P|3P)\s*(\d+A)\s*(接地|アース|E)?',
        ],
        '配管ヒーター': [
            r'(配管ヒーター|ヒーティングケーブル|ヒートトレース|自己制御型)',
            r'(SRF|SRL|BTV|QTVR)\s*[-]?\s*(\d+)',
        ],
        '監視制御': [
            r'(監視制御|SCADA|PLC|シーケンサ|リモートI/O|中央監視)',
            r'(計装|変換器|トランスミッタ|信号変換)',
        ],
        '幹線': [
            r'(高圧幹線|低圧幹線|幹線|バスダクト)',
            r'(CV[T]?)\s*(\d+)\s*(?:sq|mm)?\s*[-x×]\s*(\d+)C',
        ],
        '接地': [
            r'(接地工事|アース|接地極|接地線|A種|B種|C種|D種)\s*(接地)?',
            r'(IV|GV)\s*(\d+)\s*(?:sq|mm)?\s*(緑|G)',
        ],
    }

    extracted = []
    seen = set()

    for page_num, text in page_texts.items():
        if not text:
            continue

        for category, pattern_list in patterns.items():
            for pattern in pattern_list:
                matches = re.finditer(pattern, text, re.IGNORECASE)
                for match in matches:
                    full_match = match.group(0).strip()
                    if len(full_match) < 2:
                        continue

                    # Build a unique key to avoid duplicates
                    key = f"{category}:{full_match}"
                    if key in seen:
                        # Increment quantity instead
                        for item in extracted:
                            if item.get('_key') == key:
                                item['quantity'] = item.get('quantity', 1) + 1
                                break
                        continue
                    seen.add(key)

                    # Build item
                    item = {
                        '_key': key,
                        'page_number': page_num,
                        'material_name': _build_material_name(category, match),
                        'spec': _build_spec(match),
                        'quantity': 1,
                        'unit': _get_default_unit(category),
                        'construction_method': '',
                        'field_category': category,
                        'confidence': 0.5,
                        'match_type': 'ai_detected',
                        'reason': f'AI図面解析: ページ{page_num}で検出',
                    }
                    extracted.append(item)

    # Clean up internal keys
    for item in extracted:
        item.pop('_key', None)

    return extracted


def _build_material_name(category, match):
    """Build material name from regex match"""
    groups = [g for g in match.groups() if g]
    if category == 'ケーブル':
        return groups[0] if groups else category
    elif category == '電線管':
        return groups[0] if groups else '電線管'
    return groups[0] if groups else category


def _build_spec(match):
    """Build specification from regex match"""
    groups = [g for g in match.groups() if g]
    if len(groups) > 1:
        return ' '.join(groups[1:])
    return ''


def _get_default_unit(category):
    """Get default unit for a material category"""
    unit_map = {
        'ケーブル': 'm',
        '電線管': 'm',
        'ケーブルラック': 'm',
        '配電盤・分電盤': '面',
        '照明器具': '台',
        'コンセント・スイッチ': '個',
        '配管ヒーター': 'm',
        '監視制御': '式',
        '幹線': 'm',
        '接地': '式',
    }
    return unit_map.get(category, '式')


@app.route('/projects/<int:project_id>/blueprint-items', methods=['POST'])
@login_required
def save_blueprint_items(project_id):
    """Save blueprint extracted items"""
    try:
        db = get_db()
        cursor = db.cursor()

        data = request.get_json()
        items = data.get('items', [])
        file_id = data.get('file_id')

        if not file_id:
            return jsonify({'error': 'file_id is required'}), 400

        saved_count = 0
        for item in items:
            cursor.execute(
                '''INSERT INTO blueprint_items
                   (project_id, file_id, page_number, material_name, spec, quantity, unit,
                    construction_method, field_category, confidence, match_type, reason, master_id, is_adopted)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                (project_id, file_id, item.get('page_number', 1),
                 item.get('material_name', ''), item.get('spec', ''),
                 float(item.get('quantity', 0)), item.get('unit', ''),
                 item.get('construction_method', ''), item.get('field_category', ''),
                 float(item.get('confidence', 0)), item.get('match_type', 'manual'),
                 item.get('reason', ''), item.get('master_id'),
                 1 if item.get('is_adopted', True) else 0)
            )
            saved_count += 1

        db.commit()

        return jsonify({
            'success': True,
            'message': f'{saved_count}件の拾い出しアイテムを保存しました',
            'saved_count': saved_count
        }), 200

    except Exception as e:
        add_error_log(current_user.id, 'SAVE_BLUEPRINT_ITEMS_ERROR', str(e), str(e), request.url)
        return jsonify({'error': f'保存エラー: {str(e)}'}), 500


@app.route('/projects/<int:project_id>/blueprint-items/to-material-list', methods=['POST'])
@login_required
def blueprint_items_to_material_list(project_id):
    """Convert blueprint items to material list for estimation"""
    try:
        db = get_db()
        cursor = db.cursor()

        data = request.get_json()
        file_id = data.get('file_id')

        # Get adopted blueprint items
        cursor.execute(
            '''SELECT material_name, spec, quantity, unit, construction_method, field_category
               FROM blueprint_items
               WHERE project_id = ? AND file_id = ? AND is_adopted = 1
               ORDER BY id''',
            (project_id, file_id)
        )
        items = cursor.fetchall()

        if not items:
            return jsonify({'error': '採用済みアイテムがありません'}), 400

        # Get max row_no in material_list
        cursor.execute('SELECT COALESCE(MAX(row_no), 0) FROM material_list WHERE project_id = ?', (project_id,))
        max_row = cursor.fetchone()[0]

        count = 0
        for i, item in enumerate(items, 1):
            cursor.execute(
                '''INSERT INTO material_list
                   (project_id, row_no, material_name, spec, quantity, unit, construction_method, field_category, drawing_ref)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                (project_id, max_row + i, item[0], item[1], item[2], item[3], item[4], item[5], f'図面拾い出し')
            )
            count += 1

        db.commit()

        return jsonify({
            'success': True,
            'message': f'{count}件を材料リストに追加しました',
            'count': count
        }), 200

    except Exception as e:
        add_error_log(current_user.id, 'BLUEPRINT_TO_MATERIAL_ERROR', str(e), str(e), request.url)
        return jsonify({'error': str(e)}), 500


@app.route('/api/master-search')
@login_required
def master_search():
    """Search master data for material selection"""
    try:
        query = request.args.get('q', '').strip()
        if not query or len(query) < 1:
            return jsonify({'results': []}), 200

        db = get_db()
        cursor = db.cursor()

        # Search by name, spec, category
        search_term = f'%{query}%'
        cursor.execute(
            '''SELECT id, material_name, spec_summary, construction_method, unit,
                      composite_unit_price, field_category, source_page, category_no
               FROM estimate_master
               WHERE material_name LIKE ? OR spec_summary LIKE ? OR field_category LIKE ? OR category_no LIKE ?
               LIMIT 50''',
            (search_term, search_term, search_term, search_term)
        )

        results = []
        for row in cursor.fetchall():
            results.append({
                'id': row[0],
                'material_name': row[1],
                'spec': row[2],
                'construction_method': row[3],
                'unit': row[4],
                'composite_unit_price': float(row[5] or 0),
                'field_category': row[6],
                'source_page': row[7],
                'category_no': row[8],
            })

        return jsonify({'results': results}), 200

    except Exception as e:
        return jsonify({'error': str(e), 'results': []}), 500


# ==================== SHARED FILES ====================

@app.route('/shared-files')
@login_required
def shared_files():
    """Shared files management page"""
    try:
        db = get_db()
        cursor = db.cursor()

        project_id = request.args.get('project_id', type=int)

        # Get projects for sidebar
        if current_user.is_admin():
            cursor.execute(
                'SELECT id, name, client_name, created_at, created_by FROM projects ORDER BY created_at DESC'
            )
        else:
            cursor.execute(
                'SELECT id, name, client_name, created_at, created_by FROM projects WHERE created_by = ? ORDER BY created_at DESC',
                (current_user.id,)
            )
        projects = [dict(r) for r in cursor.fetchall()]

        # Get files
        if project_id:
            cursor.execute(
                '''SELECT sf.id, sf.project_id, p.name as project_name, sf.original_name,
                          sf.file_type, sf.file_size, u.full_name as uploaded_by_name, sf.uploaded_at
                   FROM shared_files sf
                   JOIN projects p ON sf.project_id = p.id
                   JOIN users u ON sf.uploaded_by = u.id
                   WHERE sf.project_id = ?
                   ORDER BY sf.uploaded_at DESC''',
                (project_id,)
            )
        elif current_user.is_admin():
            cursor.execute(
                '''SELECT sf.id, sf.project_id, p.name as project_name, sf.original_name,
                          sf.file_type, sf.file_size, u.full_name as uploaded_by_name, sf.uploaded_at
                   FROM shared_files sf
                   JOIN projects p ON sf.project_id = p.id
                   JOIN users u ON sf.uploaded_by = u.id
                   ORDER BY sf.uploaded_at DESC'''
            )
        else:
            cursor.execute(
                '''SELECT sf.id, sf.project_id, p.name as project_name, sf.original_name,
                          sf.file_type, sf.file_size, u.full_name as uploaded_by_name, sf.uploaded_at
                   FROM shared_files sf
                   JOIN projects p ON sf.project_id = p.id
                   JOIN users u ON sf.uploaded_by = u.id
                   WHERE p.created_by = ?
                   ORDER BY sf.uploaded_at DESC''',
                (current_user.id,)
            )
        files = [dict(r) for r in cursor.fetchall()]

        # Format file sizes
        for f in files:
            size = f.get('file_size', 0) or 0
            if size >= 1048576:
                f['file_size_display'] = f'{size / 1048576:.1f} MB'
            elif size >= 1024:
                f['file_size_display'] = f'{size / 1024:.1f} KB'
            else:
                f['file_size_display'] = f'{size} B'

        current_project = None
        if project_id:
            for p in projects:
                if p['id'] == project_id:
                    current_project = p
                    break

        return render_template(
            'shared_files.html',
            projects=projects,
            current_project=current_project,
            files=files,
            is_admin=current_user.is_admin()
        )

    except Exception as e:
        add_error_log(current_user.id, 'SHARED_FILES_ERROR', str(e), str(e), request.url)
        return render_template('error.html', error_code=500, error_message='共有ファイル読み込みエラー'), 500


@app.route('/shared-files/upload', methods=['POST'])
@login_required
def upload_shared_file():
    """Upload a shared file"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        file = request.files['file']
        sf_project_id = request.form.get('project_id', type=int)

        if not sf_project_id:
            return jsonify({'error': '案件を選択してください'}), 400

        if file.filename == '':
            return jsonify({'error': 'ファイルが選択されていません'}), 400

        original_name = file.filename
        file_ext = original_name.rsplit('.', 1)[1].lower() if '.' in original_name else ''

        # Create shared upload directory
        shared_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'shared', str(sf_project_id))
        os.makedirs(shared_dir, exist_ok=True)

        # Save file
        unique_filename = f"{uuid.uuid4().hex}.{file_ext}" if file_ext else f"{uuid.uuid4().hex}"
        file_path = os.path.join(shared_dir, unique_filename)
        file.save(file_path)

        file_size = os.path.getsize(file_path)

        # Determine file type
        type_map = {'pdf': 'PDF', 'xlsx': 'Excel', 'xls': 'Excel', 'csv': 'CSV',
                     'tsv': 'TSV', 'txt': 'テキスト', 'shd': 'SHD', 'str': 'STR'}
        file_type = type_map.get(file_ext, 'その他')

        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            '''INSERT INTO shared_files
               (project_id, original_name, stored_path, file_type, file_size, uploaded_by)
               VALUES (?,?,?,?,?,?)''',
            (sf_project_id, original_name, file_path, file_type, file_size, current_user.id)
        )
        db.commit()

        return jsonify({
            'success': True,
            'message': f'{original_name} をアップロードしました'
        }), 200

    except Exception as e:
        add_error_log(current_user.id, 'UPLOAD_SHARED_FILE_ERROR', str(e), str(e), request.url)
        return jsonify({'error': f'アップロードエラー: {str(e)}'}), 500


@app.route('/shared-files/download/<int:sf_file_id>')
@login_required
def download_shared_file(sf_file_id):
    """Download a shared file"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute(
            'SELECT stored_path, original_name FROM shared_files WHERE id = ?',
            (sf_file_id,)
        )
        file_row = cursor.fetchone()
        if not file_row:
            return jsonify({'error': 'ファイルが見つかりません'}), 404

        return send_file(file_row[0], as_attachment=True, download_name=file_row[1])

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/shared-files/delete/<int:sf_file_id>', methods=['DELETE'])
@login_required
def delete_shared_file(sf_file_id):
    """Delete a shared file (admin only)"""
    try:
        if not current_user.is_admin():
            return jsonify({'error': '管理者権限が必要です'}), 403

        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT stored_path FROM shared_files WHERE id = ?', (sf_file_id,))
        file_row = cursor.fetchone()
        if not file_row:
            return jsonify({'error': 'ファイルが見つかりません'}), 404

        # Delete file from disk
        if os.path.exists(file_row[0]):
            os.remove(file_row[0])

        cursor.execute('DELETE FROM shared_files WHERE id = ?', (sf_file_id,))
        db.commit()

        return jsonify({'success': True, 'message': 'ファイルを削除しました'}), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== AI ESTIMATION ====================

@app.route('/projects/<int:project_id>/ai-estimate', methods=['POST'])
@login_required
def ai_estimate(project_id):
    """AI-powered estimation: match materials against master and create estimate"""
    try:
        db = get_db()
        cursor = db.cursor()

        # Check project
        cursor.execute('SELECT created_by FROM projects WHERE id = ?', (project_id,))
        project = cursor.fetchone()
        if not project or (project[0] != current_user.id and not current_user.is_admin()):
            return jsonify({'error': 'アクセス権限がありません'}), 403

        # Run matching engine
        results = run_project_matching(project_id, current_user.id)

        if 'error' in results:
            return jsonify({'error': results['error']}), 400

        # Get labor unit price
        cursor.execute("SELECT setting_value FROM estimate_settings WHERE setting_key='labor_unit_price'")
        lup_row = cursor.fetchone()
        labor_unit_price = float(lup_row[0]) if lup_row else 25000

        # Calculate totals
        cursor.execute(
            '''SELECT COALESCE(SUM(amount), 0), COALESCE(SUM(productivity_total), 0), COUNT(*)
               FROM estimate_details WHERE project_id = ?''',
            (project_id,)
        )
        totals = cursor.fetchone()
        total_material = totals[0]
        total_productivity = totals[1]
        total_labor = total_productivity * labor_unit_price
        detail_count = totals[2]

        # Update project status
        cursor.execute(
            'UPDATE projects SET status = ?, updated_at = ? WHERE id = ?',
            ('estimated', datetime.utcnow(), project_id)
        )
        db.commit()

        return jsonify({
            'success': True,
            'message': f'AI見積完了: {detail_count}件の明細を作成しました',
            'results': results,
            'totals': {
                'material_cost': total_material,
                'labor_cost': total_labor,
                'productivity_total': total_productivity,
                'detail_count': detail_count,
            }
        }), 200

    except Exception as e:
        add_error_log(current_user.id, 'AI_ESTIMATE_ERROR', str(e), str(e), request.url)
        return jsonify({'error': f'AI見積エラー: {str(e)}'}), 500


# ==================== DEBUG ====================

@app.route('/debug/errors')
@login_required
def debug_errors():
    """Show recent errors as JSON for debugging"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT error_type, error_message, traceback, url, created_at FROM error_log ORDER BY created_at DESC LIMIT 20')
        errors = [{'type': r[0], 'message': r[1], 'traceback': r[2], 'url': r[3], 'time': r[4]} for r in cursor.fetchall()]
        return jsonify({'errors': errors}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/debug/project/<int:project_id>')
@login_required
def debug_project(project_id):
    """Debug project detail rendering step by step"""
    import traceback as tb
    steps = []
    try:
        db = get_db()
        cursor = db.cursor()
        steps.append('DB connected')

        cursor.execute(
            'SELECT id, name, description, client_name, status, created_at, created_by FROM projects WHERE id = ?',
            (project_id,)
        )
        project = cursor.fetchone()
        steps.append(f'Project fetched: {project is not None}')
        if not project:
            return jsonify({'steps': steps, 'error': 'Project not found'}), 404

        cursor.execute(
            '''SELECT id, file_type, original_name, file_size, uploaded_at
               FROM project_files WHERE project_id = ? ORDER BY uploaded_at DESC''',
            (project_id,)
        )
        files = cursor.fetchall()
        steps.append(f'Files fetched: {len(files)}')
        if files:
            f = files[0]
            steps.append(f'File[0] type: {type(f).__name__}, keys: {list(f.keys()) if hasattr(f, "keys") else "no keys"}')
            try:
                steps.append(f'File[0].original_name = {f["original_name"]}')
            except Exception as e2:
                steps.append(f'File[0]["original_name"] error: {e2}')

        cursor.execute(
            '''SELECT id, row_no, material_name, spec, size, quantity, unit, construction_method, field_category
               FROM material_list WHERE project_id = ? ORDER BY row_no''',
            (project_id,)
        )
        materials = cursor.fetchall()
        steps.append(f'Materials fetched: {len(materials)}')

        cursor.execute(
            'SELECT COUNT(*), COALESCE(SUM(amount),0), COALESCE(SUM(productivity_total),0) FROM estimate_details WHERE project_id = ?',
            (project_id,)
        )
        est_row = cursor.fetchone()
        estimate_count = est_row[0]
        total_amount = est_row[1]
        total_productivity = est_row[2]
        steps.append(f'Estimates: count={estimate_count}, amount={total_amount}, productivity={total_productivity}')

        cursor.execute("SELECT setting_value FROM estimate_settings WHERE setting_key='labor_unit_price'")
        lup_row = cursor.fetchone()
        labor_unit_price = float(lup_row[0]) if lup_row else 25000
        steps.append(f'Labor unit price: {labor_unit_price}')

        cursor.execute(
            '''SELECT id, material_id, candidate_rank, master_id, match_type,
               confidence, reason, is_adopted, master_name, master_spec,
               master_method, composite_unit_price, removal_productivity, source_page
            FROM match_results WHERE project_id = ? ORDER BY material_id, candidate_rank''',
            (project_id,)
        )
        match_results = cursor.fetchall()
        steps.append(f'Match results: {len(match_results)}')

        # Try rendering template
        try:
            html = render_template(
                'project_detail.html',
                project={
                    'id': project[0], 'name': project[1], 'description': project[2],
                    'client_name': project[3], 'status': project[4], 'created_at': project[5]
                },
                files=files, materials=materials,
                estimate_count=estimate_count, total_amount=total_amount,
                total_productivity=total_productivity,
                labor_unit_price=labor_unit_price, match_results=match_results
            )
            steps.append(f'Template rendered OK, length={len(html)}')
        except Exception as te:
            steps.append(f'TEMPLATE ERROR: {te}')
            steps.append(tb.format_exc())

        return jsonify({'steps': steps, 'status': 'OK'}), 200

    except Exception as e:
        steps.append(f'EXCEPTION: {e}')
        steps.append(tb.format_exc())
        return jsonify({'steps': steps, 'status': 'ERROR'}), 500


# ==================== VERSION & ERROR HANDLERS ====================

APP_VERSION = 'v1.3.0'

@app.route('/debug/version')
def debug_version():
    """Show app version for deploy verification"""
    return jsonify({'version': APP_VERSION, 'status': 'ok'}), 200


@app.route('/health')
def health_check():
    """Health check endpoint for Render and monitoring"""
    try:
        db = get_db()
        cursor = db.cursor()
        cursor.execute('SELECT 1')
        return jsonify({'status': 'healthy', 'version': APP_VERSION}), 200
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 503


@app.errorhandler(404)
def not_found(error):
    """404 error handler"""
    return render_template('error.html', error_code=404, error_message=f'ページが見つかりません (v{APP_VERSION})'), 404


@app.errorhandler(500)
def internal_error(error):
    """500 error handler"""
    import traceback
    tb_str = traceback.format_exc()
    return render_template('error.html', error_code=500,
                           error_message=f'サーバーエラーが発生しました',
                           error_details=f'{error}\n\n{tb_str}'), 500


@app.errorhandler(403)
def forbidden(error):
    """403 error handler"""
    return render_template('error.html', error_code=403, error_message='アクセスが禁止されています'), 403


# Enable DEBUG mode temporarily to show error details on error page
app.config['DEBUG'] = True


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
